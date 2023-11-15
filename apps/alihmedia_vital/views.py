from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from .models import Doc, Variety
from django.contrib import messages
from django.db.models import Q
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
from os.path import exists
from .forms import SearchDoc, InsertPdfDoc, UploadFileForm, DocAddForm
import sys
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
import roman
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import permission_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Max

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
def index(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.GET.get("folder"):
        folder = request.GET.get("folder")
        return redirect(f"/{__package__.split('.')[1]}/{folder}")
    context = {}
    context['form'] = InsertPdfDoc()
    return render(request,'alihmedia_vital/index.html', context=context)

def getdata(folder):
    d = Variety.objects.get(folder=folder)
    docs = Doc.objects.filter(variety_id__exact=d.id)
    docdata = []
    for ke, doc in enumerate(docs):
        path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], doc.variety.folder, str(doc.doc_number) + ".pdf")
        pdffound = False
        coverfilename = ""
        if exists(path):
            pdffound = True
            coverfilename = "{}_{}_{}.png".format(__package__.split('.')[1], doc.variety.folder, doc.doc_number)
        filetmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc.id}.pdf")
        pdftmpfound = False
        if exists(filetmppath):
            pdftmpfound = True

        docdata.append({
            "variety": doc.variety.name,
            "doc_number": doc.doc_number,
            "name": doc.name,
            "work_unit": doc.work_unit,
            "period": doc.period,
            "media": doc.media,
            "countstr": doc.countstr,
            "save_life": doc.save_life,
            "uuid_id": doc.uuid_id,
            "save_location": doc.save_location,
            "protect_method": doc.protect_method,
            "description": doc.description,
            "pdffound": pdffound,
            "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
            "filesize": doc.filesize,
            "pagecount": doc.page_count,
            "pdftmpfound": pdftmpfound,


        })
    return docdata
@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
def sertifikat(request):
    if not request.user.is_authenticated:
        return redirect('login')
    folder = sys._getframe().f_code.co_name
    d = Variety.objects.get(folder=folder)
    docdata = getdata(folder)
    context = {'data':docdata, "title": d.name, "folder": folder}
    return render(request,'alihmedia_vital/datalist.html', context=context)

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
def bpkb_mobil_dan_motor(request):
    if not request.user.is_authenticated:
        return redirect('login')
    folder = sys._getframe().f_code.co_name
    d = Variety.objects.get(folder=folder)
    docdata = getdata(folder)
    context = {'data':docdata, "title": d.name, "folder": folder}
    return render(request,'alihmedia_vital/datalist.html', context=context)

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
def searchdoc(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.GET.get("folder"):
        query = request.GET.get("search")
        folder = request.GET.get("folder")
        d = Variety.objects.get(folder=folder)
        if query == None or query == '':
            docs = Doc.objects.filter(variety_id__exact=d.id)
        else:
            docs = Doc.objects.filter(Q(variety_id__exact=d.id) & (Q(name__icontains=query)  | Q(work_unit__icontains=query) | Q(period__exact=query ) | Q(save_location__icontains=query)))
        if not docs:
            messages.info(request, "Data tidak ditemukan")
        docdata = []
        for ke, doc in enumerate(docs):
            path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], doc.variety.folder, str(doc.doc_number) + ".pdf")
            pdffound = False
            coverfilename = ""
            if exists(path):
                pdffound = True
                coverfilename = "{}_{}_{}.png".format(__package__.split('.')[1], doc.variety.folder, doc.doc_number)

            docdata.append({
                "variety": doc.variety.name,
                "doc_number": doc.doc_number,
                "name": doc.name,
                "work_unit": doc.work_unit,
                "period": doc.period,
                "media": doc.media,
                "countstr": doc.countstr,
                "save_life": doc.save_life,
                "uuid_id": doc.uuid_id,
                "save_location": doc.save_location,
                "protect_method": doc.protect_method,
                "description": doc.description,
                "pdffound": pdffound,
                "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
                "filesize": doc.filesize,
                "pagecount": doc.page_count,
            })
        context = {'data':docdata, 'form': SearchDoc(), 'folder':folder, 'query':query}
        return render(request,'alihmedia_vital/searchdoc.html', context=context)

    context = {}
    context['form'] = SearchDoc()
    return render(request,'alihmedia_vital/searchdoc.html', context=context)

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all())
@csrf_exempt
def pdfupload(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')
    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.variety.folder
    doc_id = doc.id
    pdfpath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], str(doc.doc_number) + ".pdf")
    tmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc_id}.pdf")
    if exists(pdfpath):
        messages.info(request, "File sudah ada")
        return redirect(f"/{__package__.split('.')[1]}/{folder}")


    if request.method == 'POST' and request.FILES['filepath'] and not exists(pdfpath):
        upload = request.FILES['filepath']
        fss = FileSystemStorage()
        if exists(tmppath):
            os.remove(tmppath)
        fss.save(tmppath, upload)
        messages.info(request, "File berhasil diupload, akan segera diproses")
        # time.sleep(2)
        return redirect(f"/{__package__.split('.')[1]}/{folder}")

    context = {}
    context['form'] = UploadFileForm(initial={'uuid_id': uuid_id})
    context['isexist'] = exists(tmppath)
    context['data'] = doc

    # context['url'] = url
    return render(request,'alihmedia_vital/pdfupload.html', context=context)

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
def pdfdownload(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')

    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.variety.folder
    doc_number = doc.doc_number
    
    path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc_number) + ".pdf")
    if exists(path):
        filename = f"{__package__.split('.')[1]}_{folder}_{doc_number}.pdf"
        with open(path, 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline;filename={filename}'
            return response
    raise Http404

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all())
@csrf_exempt
def delete(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')
    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.variety.folder
    doc_id = doc.id
    pdfpath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.doc_number) + ".pdf")
    coverfilename = "{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.doc_number)
    if request.method == 'POST':
        if exists(pdfpath):
            os.remove(pdfpath)
            coverfilename = "{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.doc_number)
            if exists(os.path.join(settings.COVER_LOCATION, coverfilename)):
                os.remove(os.path.join(settings.COVER_LOCATION, coverfilename))
            tmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc_id}.pdf")
            if exists(tmppath):
                os.remove(tmppath)
        doc.delete()     
        messages.info(request, "Berhasil dihapus")
        return redirect(f"/{__package__.split('.')[1]}/{folder}")
    context = {}
    context['uuid_id'] = uuid_id
    context['isexist'] = exists(pdfpath)
    context['data'] = doc
    context["coverfilepath"] =  os.path.join(settings.COVER_URL, coverfilename)

    # context['url'] = url
    return render(request,'alihmedia_vital/delete.html', context=context)
@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all())
@csrf_exempt
def pdfremove(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')
    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.variety.folder
    doc_id = doc.id
    pdfpath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.doc_number) + ".pdf")
    if not exists(pdfpath):
        messages.info(request, "File tidak ada")
        return redirect(f"/{__package__.split('.')[1]}/{folder}")
    coverfilename = "{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.doc_number)
    if request.method == 'POST':
        if exists(pdfpath):
            os.remove(pdfpath)
            coverfilename = "{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.doc_number)
            if exists(os.path.join(settings.COVER_LOCATION, coverfilename)):
                os.remove(os.path.join(settings.COVER_LOCATION, coverfilename))
            tmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc_id}.pdf")
            if exists(tmppath):
                os.remove(tmppath)
            doc.filesize = None
            doc.page_count = None
            doc.save()     
            messages.info(request, "Berhasil dihapus")
            return redirect(f"/{__package__.split('.')[1]}/{folder}")
        else:
            messages.info(request, "File tidak ada")
    context = {}
    context['uuid_id'] = uuid_id
    context['isexist'] = exists(pdfpath)
    context['data'] = doc
    context["coverfilepath"] =  os.path.join(settings.COVER_URL, coverfilename)

    # context['url'] = url
    return render(request,'alihmedia_vital/pdfremove.html', context=context)

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
@csrf_exempt
def update(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')
    doc = Doc.objects.get(uuid_id=uuid_id)
    updateForm=DocAddForm(instance=doc)
    if request.method=='POST':
        docAdd=DocAddForm(request.POST,instance=doc)
        if docAdd.is_valid():
            docAdd.save()
            messages.success(request, 'Data has been updated.')
            next = request.POST.get('next', '/')
            return redirect(next)
    return render(request,'alihmedia_vital/update.html',{
        'form':updateForm,
        'item':doc})

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
@csrf_exempt
def add(request):
    if not request.user.is_authenticated:
        return redirect('login')
    addForm=DocAddForm()
    if request.method=='POST':
        docAdd=DocAddForm(request.POST)
        folder = request.POST.get('folder')
        variety = Variety.objects.get(folder=folder)
        docmax = Doc.objects.filter(variety=variety).aggregate(Max('doc_number'))
        if docAdd.is_valid():
            instance = docAdd.save(commit=False)
            instance.variety_id = variety.id
            instance.doc_number = docmax['doc_number__max'] + 1
            instance.save()
            messages.success(request, 'Data has been added.')
            next = request.POST.get('next', '/')
            return redirect(next)
    return render(request,'alihmedia_vital/add.html',{
        'form':addForm})

def create_xls(datalist):
    wb = Workbook()
    wb.create_sheet("CONFIG")
    sheet = wb["CONFIG"]
    sheet['A1'].value = os.path.join("D:", "media") + "\\"

    sheet = wb.active
    sheet.title = "DATA INAKTIF"
    sheet.column_dimensions['A'].width = 4.1
    sheet.column_dimensions['B'].width = 37
    sheet.column_dimensions['C'].width = 11.3
    sheet.column_dimensions['D'].width = 8.8
    sheet.column_dimensions['E'].width = 9.5
    sheet.column_dimensions['F'].width = 12.8
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['H'].width = 23.5 
    sheet.column_dimensions['I'].width = 15.5 
    sheet.column_dimensions['J'].width = 14.6
    sheet.column_dimensions['K'].width = 8.8
    sheet.merge_cells('A1:K1')
    sheet['A1'] = "DAFTAR ARSIP VITAL"
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].font = Font(name='Arial', size=12, bold=True)
    centervh = Alignment(horizontal='center', vertical='center', wrap_text=True)
    centerv = Alignment(vertical='center', wrap_text=True)

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    thin_border2 = Border(bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
    thin_border3 = Border(right=Side(style='thin'),left=Side(style='thin'))
    font_style1 = Font(name='Arial', size=8.5, bold=True)
    font_style2 = Font(name='Arial', size=8.5)
    font_style3 = Font(name='Arial', size=8.5, italic=True)
    color1 = PatternFill(start_color="c6d5f7", fill_type = "solid")


    sheet.row_dimensions[7].height = 28
    # for cell in sheet["7:7"]:
    #     cell.alignment = centervh
    #     cell.font = font_style1
    headers = ("NO", "JENIS ARSIP", "UNIT KERJA", "KURUN WAKTU", "MEDIA", "JUMLAH", "JANGKA SIMPAN", "LOKASI SIMPAN", "METODE PERLINDUNGAN", "KETERANGAN", "ACTION")
    for i in headers:
        sheet.cell(row=7, column=headers.index(i)+1).value = i
        sheet.cell(row=7, column=headers.index(i)+1).border = thin_border
        sheet.cell(row=7, column=headers.index(i)+1).alignment = centervh
        sheet.cell(row=7, column=headers.index(i)+1).font = font_style1
    row = 8
    for idx1, d in enumerate(datalist):
        sheet.cell(row=row, column=1).value = roman.toRoman(idx1+1)
        sheet.cell(row=row, column=1).alignment = centervh
        sheet.cell(row=row, column=1).font = font_style1
        sheet.cell(row=row, column=2).value = d['variety']
        sheet.cell(row=row, column=2).font = font_style1
        
        for cell in sheet["{}:{}".format(row,row) ]:
            cell.border = thin_border3
        
        row += 1
        for idx2, data in enumerate(d['data']):
            sheet.cell(row=row, column=1).value = idx2+1
            sheet.cell(row=row, column=1).alignment = centerv
            sheet.cell(row=row, column=1).border = thin_border2
            sheet.cell(row=row, column=1).font = font_style2
            
            sheet.cell(row=row, column=2).value = data['name']
            sheet.cell(row=row, column=2).alignment = centerv
            sheet.cell(row=row, column=2).border = thin_border2
            sheet.cell(row=row, column=2).font = font_style2

            sheet.cell(row=row, column=3).value = data['work_unit']
            sheet.cell(row=row, column=3).alignment = centervh
            sheet.cell(row=row, column=3).border = thin_border2
            sheet.cell(row=row, column=3).font = font_style2

            sheet.cell(row=row, column=4).value = data['period']
            sheet.cell(row=row, column=4).alignment = centervh
            sheet.cell(row=row, column=4).border = thin_border2
            sheet.cell(row=row, column=4).font = font_style2

            sheet.cell(row=row, column=5).value = data['media']
            sheet.cell(row=row, column=5).alignment = centervh
            sheet.cell(row=row, column=5).border = thin_border2
            sheet.cell(row=row, column=5).font = font_style2

            sheet.cell(row=row, column=6).value = data['countstr']
            sheet.cell(row=row, column=6).alignment = centervh
            sheet.cell(row=row, column=6).border = thin_border2
            sheet.cell(row=row, column=6).font = font_style2

            sheet.cell(row=row, column=7).value = data['save_life']
            sheet.cell(row=row, column=7).alignment = centervh
            sheet.cell(row=row, column=7).border = thin_border2
            sheet.cell(row=row, column=7).font = font_style2
            
            sheet.cell(row=row, column=8).value = data['save_location']
            sheet.cell(row=row, column=8).alignment = centerv
            sheet.cell(row=row, column=8).border = thin_border2
            sheet.cell(row=row, column=8).font = font_style2

            sheet.cell(row=row, column=9).value = data['protect_method']
            sheet.cell(row=row, column=9).alignment = centervh
            sheet.cell(row=row, column=9).border = thin_border2
            sheet.cell(row=row, column=9).font = font_style3

            sheet.cell(row=row, column=10).value = data['description']
            sheet.cell(row=row, column=10).alignment = centervh
            sheet.cell(row=row, column=10).border = thin_border2
            sheet.cell(row=row, column=10).font = font_style2
            
            sheet.cell(row=row, column=11).value = ""
            sheet.cell(row=row, column=11).alignment = centervh
            sheet.cell(row=row, column=11).border = thin_border2
            sheet.cell(row=row, column=11).font = font_style2
            if data['filesize'] is not None:
                filelocation = os.path.join(__package__.split('.')[1], d['folder'], str(data['doc_number']) + ".pdf")
                sheet.cell(row=row, column=11).fill = color1
                sheet.cell(row=row, column=2).fill = color1
                sheet.cell(row=row, column=11).value = '=HYPERLINK(CONCATENATE(CONFIG!A1, "{}")'.format(filelocation) + ', "LIHAT")'
                sheet.cell(row=row, column=11).font = Font(underline='single', bold=True, color="96251b")

            row += 1
    
    
    return wb

@user_passes_test(lambda user: Group.objects.get(name='BMN') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all(), login_url='alihmedia_vital_authorization_rejected')
def export(request):
    if not request.user.is_authenticated:
        return redirect('login')
    varieties = Variety.objects.all()
    datalist = []
    for vr in varieties:
        mdict = {}
        data = getdata(vr.folder)
        mdict["variety"] = vr.name
        mdict["folder"] = vr.folder
        mdict["data"] = data
        datalist.append(mdict)

    filename = f"data_{__package__.split('.')[1]}.xlsx"
    # return HttpResponse(json.dumps(datalist, default=str), content_type="application/json")

    wb = create_xls(datalist)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)    
    wb.save(response)
    return response


def authorization_rejected(request):
    return render(request=request, template_name='page_404.html', context={'message':'Otorisasi Ditolak'})

from django.shortcuts import render
import os
from django.shortcuts import render, get_object_or_404, redirect, get_list_or_404
from django.http import HttpResponse, Http404, JsonResponse, FileResponse
from .models import Year, Box, Bundle, Item, Customer, Trans, TransDetail
from django.views.decorators.csrf import csrf_exempt
import json
from .forms import YearForm, BoxForm, BundleForm, ItemForm, CustomerForm, TransForm, AddTransDetailForm, EditTransDetailForm, SearchItemForm, SearchBundleForm, SearchDocByYear
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from reportlab.lib.pagesizes import A4, A5
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io
from reportlab.platypus import SimpleDocTemplate, Paragraph, PageTemplate, BaseDocTemplate, Frame, Spacer
from reportlab.lib.units import inch, mm
from reportlab.platypus.tables import Table,TableStyle,colors
from datetime import datetime, timedelta
from django.template.defaultfilters import slugify
from django.db.models import Q, Max, Count, Sum
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
from django.conf import settings
from os.path import exists
from django.views.decorators.http import require_http_methods
from reportlab_qrcode import QRCodeImage
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import permission_required, user_passes_test
from django.core.files.storage import FileSystemStorage
import fitz
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from urllib.parse import unquote, quote, unquote_plus, quote_plus
from django.utils import timezone
# Create your views here.

def is_user_in_group(user, group_name):
    try:
        group = Group.objects.get(name=group_name)
        return group in user.groups.all()
    except Group.DoesNotExist:
        return False
    
@csrf_exempt
def year_list(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'arsip_tata/year_list.html', {
        'years': Year.objects.all(),
    })
    
def show_year(request):
    if not request.user.is_authenticated:
        return redirect('login')
    context = {
    }
    return render(request=request, template_name='arsip_tata/show_year.html', context=context)
    
def add_year(request):
    if request.method == "POST":
        form = YearForm(request.POST)
        if form.is_valid():
            year = form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "yearListChanged": None,
                        "showMessage": f"{year.yeardate} added."
                    })
                })
    else:
        form = YearForm()
    return render(request, 'arsip_tata/year_form.html', {
        'form': form,
        'module': 'Tambah Data'
    })

def edit_year(request, pk):
    year = get_object_or_404(Year, pk=pk)
    # return HttpResponse(year.id)
    if request.method == "POST":
        form = YearForm(request.POST, instance=year)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "yearListChanged": None,
                        "showMessage": f"{year.yeardate} updated."
                    })
                }
            )
    else:
        form = YearForm(instance=year)
    return render(request, 'arsip_tata/year_form.html', {
        'form': form,
        'year': year,
        'module': 'Edit Data'
    })

@ require_POST
def remove_year(request, pk):
    year = get_object_or_404(Year, pk=pk)
    year.delete()
    return HttpResponse(
        status=204,
        headers={
            'HX-Trigger': json.dumps({
                "yearListChanged": None,
                "showMessage": f"{year.yeardate} deleted."
            })
        })

def show_boxes(request, year):
    if not request.user.is_authenticated:
        return redirect('login')
    yearint = year
    year = Year.objects.get(yeardate=year)
    if request.GET.get("page"):
        page = request.GET.get("page")
    else:
        page = 1
    if request.GET.get("search"):
        search = request.GET.get("search")
    else:
        search = "None"
    synccount = Item.objects.filter(issync=False, yeardate=yearint)    
    context = {
        'year_id': year.id,
        'year_date': year.yeardate,
        'page': page,
        'search': quote_plus(search),
        'synccount': synccount.count()
    }
    return render(request=request, template_name='arsip_tata/show_box.html', context=context)

def statistics(request):
    if not request.user.is_authenticated:
        return redirect('login')
    colorlist = []
    totaldoc = Item.objects.count()
    uploaded = Item.objects.filter(page_count__isnull=True).count()
    notuploaded = totaldoc - uploaded
    procuploaded = uploaded / totaldoc * 100
    procnotuploaded = notuploaded / totaldoc * 100
    colorlist.append("rgba(112, 185, 239, 1)")
    colorlist.append("rgba(244, 204, 204, 1)")
    num_of_dates = 30
    start = datetime.now()
    date_list = [start.date() - timedelta(days=x) for x in range(num_of_dates)]
    date_list.sort()
    docscan = []
    doccolor = []
    docdate = []
    # print(date_list)
    docscanmax = 100
    for d in date_list:
        pages = 0
        where = ["YEAR(uploaded_date)=%(year)s AND MONTH(uploaded_date)=%(month)s AND DAY(uploaded_date)=%(day)s" % {"year":d.year, "month": d.month, "day": d.day}]
        items = Item.objects.filter(page_count__isnull=False).extra(where=where)
        # print(d.strftime('%d-%m-%Y'))
        # print(items)
        for item in items:
            pages += item.page_count
        docdate.append(d.strftime('%d-%m-%Y'))
        docscan.append(pages)
        doccolor.append("rgba(244, 204, 204, 1)")
        if pages > docscanmax:
            docscanmax = pages
    docentry = []
    docentrycolor = []
    docentrydate = []
    docentrymax = 100
    for d in date_list:
        # pages = 0
        where = ["YEAR(created_date)=%(year)s AND MONTH(created_date)=%(month)s AND DAY(created_date)=%(day)s" % {"year":d.year, "month": d.month, "day": d.day}]
        doccount = Item.objects.extra(where=where).count()
        docentrydate.append(d.strftime('%d-%m-%Y'))
        docentry.append(doccount)
        docentrycolor.append("rgba(112, 185, 239, 1)")
        if doccount > docentrymax:
            docentrymax = doccount

    # print(docscanmax, docentrymax)
    context = {
        "totaldoc": totaldoc,
        "uploaded": uploaded,
        "notuploaded": notuploaded,
        "procuploaded": f"{procuploaded:.3f}",
        "procnotuploaded": f"{procnotuploaded:.3f}",
        "colorlist": colorlist,
        "docdate": docdate,
        "docscan": docscan,
        "doccolor": doccolor,
        "docentry": docentry,
        "docentrycolor": docentrycolor,
        "docentrydate": docentrydate,
        "docscanmax": docscanmax + (docscanmax * 0.2),
        "docentrymax": docentrymax + (docentrymax * 0.2),

    }
    # print(maxcount)
    return render(request=request, template_name='arsip_tata/statistics.html', context=context)

def statistics_year(request, year):
    if not request.user.is_authenticated:
        return redirect('login')
    
    data = Item.objects.filter(created_by__isnull=False, yeardate=year).values("created_by").annotate(count=Count('created_by'))
    for idx, rec in enumerate(data):
        username = User.objects.get(id=rec['created_by']).username
        data[idx]['username'] = username
    userlist = []
    countlist = []
    colorlist= []
    maxcount = 0
    for dt in data:
        if dt["count"] > maxcount:
            maxcount = dt["count"]
        userlist.append(dt['username'])
        countlist.append(dt['count'])
        colorlist.append("rgba(112, 185, 239, 1)")



    context = {
        "userlist": userlist,
        "countlist": countlist,
        "colorlist": colorlist,
        "maxcount": maxcount + (maxcount*0.2),

    }
    # print(maxcount)
    return render(request=request, template_name='arsip_tata/statistics_year.html', context=context)

def statistics_year_digital(request, year):
    if not request.user.is_authenticated:
        return redirect('login')
    

    data = Item.objects.filter(uploaded_by__isnull=False, uploaded_date__year=year, page_count__isnull=False).values("uploaded_by").annotate(count=Sum('page_count'))
    for idx, rec in enumerate(data):
        username = User.objects.get(id=rec['uploaded_by']).username
        data[idx]['username'] = username
    userlist_upload = []
    countlist_upload = []
    colorlist_upload = []
    maxcount_upload = 0
    for dt in data:
        if dt["count"] > maxcount_upload:
            maxcount_upload = dt["count"]
        userlist_upload.append(dt['username'])
        countlist_upload.append(dt['count'])
        colorlist_upload.append("rgba(244, 204, 204, 1)")


    context = {
        "userlist_upload": userlist_upload,
        "countlist_upload": countlist_upload,
        "colorlist_upload": colorlist_upload,
        "maxcount_upload": maxcount_upload + (maxcount_upload*0.2)

    }
    # print(maxcount)
    return render(request=request, template_name='arsip_tata/statistics_year_digital.html', context=context)

def show_boxes_old(request, year):
    if not request.user.is_authenticated:
        return redirect('login')
    year = Year.objects.get(yeardate=year)
    context = {
        'year_id': year.id,
        'year_date': year.yeardate,
    }
    return render(request=request, template_name='arsip_tata/show_box.html', context=context)

@csrf_exempt
def box_list(request, year_id, page, search):
    if not request.user.is_authenticated:
        return redirect('login')
    isadmin = is_user_in_group(request.user, 'admin')
    result = {}
    if search == "None":
        boxes = Box.objects.filter(year_id=year_id).order_by("id")
    else:
        boxes = Box.objects.filter(bundle__description__icontains=unquote_plus(search))
    # print(boxes)
    paginator = Paginator(boxes, 50)
    
    try:
        boxes = paginator.page(page)
    except PageNotAnInteger:
        boxes = paginator.page(1)
    except EmptyPage:
        boxes = paginator.page(paginator.num_pages)
    # breakpoint()
    recs = []
    for box in boxes:
        bundles = Bundle.objects.filter(box_id=box.id).all()
        bundle_numbers = []
        item_numbers = []
        bundle_years = []
        for bundle in bundles:
            bundle_numbers.append(str(bundle.bundle_number))
            bundle_years.append(str(bundle.year_bundle))
            items = Item.objects.filter(bundle_id=bundle.id).all()
            for item in items:
                item_numbers.append(item.item_number)
        bundle_numbers.sort()
        item_numbers.sort()
        bundle_years.sort()
        
        minitem = ""
        maxitem = ""
        if len(item_numbers) != 0:
            minitem = str(item_numbers[0])
            maxitem = str(item_numbers[-1])
        recs.append({"pk": box.pk, "yeardate": box.year.yeardate, "box_number": box.box_number, "bundle_number": ", ".join(bundle_numbers),  "item_number": " - ".join([minitem, maxitem]), "year_bundle": ", ".join(list(set(bundle_years))), "notes": box.notes, "itemcount": len(item_numbers), "token": box.token, "isgen": box.isgen})
    result['data'] = recs
    
    result['has_other_pages'] = boxes.has_other_pages()
    try:
        result['has_previous'] = boxes.has_previous()
        result['previous_page_number'] = boxes.previous_page_number()
    except:
        result['previous_page_number'] = False
    
    result['number'] = boxes.number
    result['page_range'] = boxes.paginator.page_range
    try:
        result['has_next'] = boxes.has_next()
        result['next_page_number'] = boxes.next_page_number()
    except:
        result['next_page_number'] = False
    result['num_pages'] = boxes.paginator.num_pages    
    return render(request, 'arsip_tata/box_list2.html', {
        'boxes': result,
        'form': SearchBundleForm(),
        'search': search,
        'isadmin': isadmin
    })

@csrf_exempt
def box_list_old(request, year_id):
    if not request.user.is_authenticated:
        return redirect('login')
    result = []
    boxes = Box.objects.filter(year_id=year_id)

    for box in boxes:
        bundles = Bundle.objects.filter(box_id=box.id).all()
        bundle_numbers = []
        item_numbers = []
        bundle_years = []
        for bundle in bundles:
            bundle_numbers.append(str(bundle.bundle_number))
            bundle_years.append(str(bundle.year_bundle))
            items = Item.objects.filter(bundle_id=bundle.id).all()
            for item in items:
                item_numbers.append(item.item_number)
        bundle_numbers.sort()
        item_numbers.sort()
        bundle_years.sort()
        
        minitem = ""
        maxitem = ""
        if len(item_numbers) != 0:
            minitem = str(item_numbers[0])
            maxitem = str(item_numbers[-1])
        result.append({"pk": box.pk, "yeardate": box.year.yeardate, "box_number": box.box_number, "bundle_number": ", ".join(bundle_numbers),  "item_number": " - ".join([minitem, maxitem]), "year_bundle": ", ".join(list(set(bundle_years))), "notes": box.notes, "itemcount": len(item_numbers), "token": box.token})
    return render(request, 'arsip_tata/box_list.html', {
        'boxes': result,
    })

@user_passes_test(lambda user: Group.objects.get(name='admin') in user.groups.all())
def add_box(request, year_id):
    if request.method == "POST":
        form = BoxForm(request.POST)
        if form.is_valid():
            box = form.save(commit=False)
            year = Year.objects.get(id=year_id)
            box.year_id = year_id
            box.yeardate = year.yeardate
            box.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "boxListChanged": None,
                        "showMessage": f"{box.box_number} added."
                    })
                })
    else:
        year = Year.objects.get(id=year_id)
        checkyear = Year.objects.filter(yeardate__gt=year.yeardate)
        if checkyear.count() > 0:
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "boxListChanged": None,
                        "showMessage": f"Tahun penataan {year.yeardate} sudah dikunci"
                    })
                })
            
        form = BoxForm(initial={'yeardate': year.yeardate})
    return render(request, 'arsip_tata/box_form.html', {
        'form': form,
    })

@user_passes_test(lambda user: Group.objects.get(name='admin') in user.groups.all())
def edit_box(request, pk):
    box = get_object_or_404(Box, pk=pk)
    if request.method == "POST":
        form = BoxForm(request.POST, instance=box)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "boxListChanged": None,
                        "showMessage": f"{box.box_number} updated."
                    })
                }
            )
    else:
        form = BoxForm(instance=box)
    return render(request, 'arsip_tata/box_form.html', {
        'form': form,
        'box': box,
        'module': 'Edit Data'
    })

# @ require_POST
@require_http_methods(['DELETE'])
def remove_box(request, pk):
    box = get_object_or_404(Box, pk=pk)
    box.delete()
    return HttpResponse(
        status=204,
        headers={
            'HX-Trigger': json.dumps({
                "boxListChanged": None,
                "showMessage": f"{box.box_number} deleted."
            })
        })

def show_bundles(request, year_date, box_number):
    if not request.user.is_authenticated:
        return redirect('login')
    year = Year.objects.get(yeardate=year_date)
    box = Box.objects.filter(year_id=year.id, box_number=box_number).first()
    
    context = {
        'box_id': box.id,
        'box_number': box_number,
        'year_date': year_date,
        'notes': box.notes
    }
    return render(request=request, template_name='arsip_tata/show_bundle.html', context=context)

@csrf_exempt
def bundle_list(request, box_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    bundles = Bundle.objects.filter(box_id=box_id).order_by("bundle_number")
    # print(bundles[0].syncstatus)
    return render(request, 'arsip_tata/bundle_list.html', {
        'bundles': bundles,
    })

# @user_passes_test(lambda user: Group.objects.get(name='admin') in user.groups.all())
def add_bundle(request, box_id):
    if request.method == "POST":
        form = BundleForm(request.POST)
        if form.is_valid():
            bundle = form.save(commit=False)
            bundle.box_id = box_id
            bundle.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "bundleListChanged": None,
                        "showMessage": f"{bundle.bundle_number} added."
                    })
                })
    else:
        
        box = Box.objects.get(id=box_id)
        if box.token != None:
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "bundleListChanged": None,
                        "showMessage": "Tidak boleh tambah berkas karena nomor sudah di generate"
                    })
                })
            
        
        try:
            latest_bundle_number = Bundle.objects.filter(box__yeardate=box.yeardate).latest('bundle_number').bundle_number + 1
        except:
            latest_bundle_number = 1

        form = BundleForm(initial={'yeardate': box.yeardate, 'box':  Box.objects.first().id, 'bundle_number': latest_bundle_number})
    return render(request, 'arsip_tata/bundle_form.html', {
        'form': form,
    })

def edit_bundle(request, pk):
    bundle = get_object_or_404(Bundle, pk=pk)
    if request.method == "POST":
        form = BundleForm(request.POST, instance=bundle)
        if form.is_valid():
            bundle = form.save(commit=False)
            bundle.syncstatus = 1
            bundle.save()
            if bundle.box.isgen:
                Item.objects.filter(bundle_id=pk).update(issync=False)
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "bundleListChanged": None,
                        "showMessage": f"{bundle.bundle_number} updated."
                    })
                }
            )
    else:
        form = BundleForm(instance=bundle)
    return render(request, 'arsip_tata/bundle_form.html', {
        'form': form,
        'bundle': bundle,
        'module': 'Edit Data'
    })

@require_http_methods(['DELETE'])
def remove_bundle(request, pk):
    bundle = get_object_or_404(Bundle, pk=pk)
    bundle.delete()
    return HttpResponse(
        status=204,
        headers={
            'HX-Trigger': json.dumps({
                "bundleListChanged": None,
                "showMessage": f"{bundle.bundle_number} deleted."
            })
        })

def show_items(request, bundle_id):
    if not request.user.is_authenticated:
        return redirect('login')
    bundle = Bundle.objects.get(id=bundle_id)
    context = {
        'bundle': bundle,
        # 'year_date': year_date
    }
    return render(request=request, template_name='arsip_tata/show_item.html', context=context)

@csrf_exempt
def item_list(request, bundle_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    items = Item.objects.filter(bundle_id=bundle_id).order_by('item_number')
    return render(request, 'arsip_tata/item_list.html', {
        'items': items,
        'coverpath': settings.COVER_URL + __package__.split('.')[1] + "_",
    })

def item_download_pdf(request, pk):
    
    if not request.user.is_authenticated:
        return redirect('login')

    item = Item.objects.get(id=pk)
    
    path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], str(item.yeardate), item.codegen + ".pdf")
    # print(path)
    if exists(path):
        doc = fitz.open(path)
        # for i in range(0, len(doc)):
        #     try:
        #         page = doc[i]
        #         tw = fitz.TextWriter(page.rect, opacity=0.3)
        #         tw.append((50, 100), "COPY")
        #         page.clean_contents()
        #         page.write_text(rect=page.rect, writers=tw)
        #     except:
        #         pass

        doc.save("tmp.pdf")            
            
        filename = f"{__package__.split('.')[1]}_{item.codegen}.pdf"
        # with open(path, 'rb') as pdf:
        with open("tmp.pdf", 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment;filename={filename}'
            return response
    raise Http404

def add_item(request, bundle_id):
    if request.method == "POST":
        # form = ItemForm(request.POST)
        form = ItemForm(request.POST or None, request.FILES or None)
        print(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.bundle_id = bundle_id
            item.total = item.copy + item.original
            item.bundle_id = bundle_id
            item.created_by = request.user
            item.created_date = datetime.now()
            # bundle = Bundle.objects.get(id=bundle_id)
            # item.codegen = "-".join([str(item.yeardate), str(bundle.box.box_number), str(bundle.bundle_number), str(item.item_number)])
            item.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "itemListChanged": None,
                        "showMessage": f"{item.item_number} added."
                    })
                })
    else:
        bundle = Bundle.objects.get(id=bundle_id)
        if bundle.box.token != None:
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "itemListChanged": None,
                        "showMessage": "Tidak boleh tambah berkas karena nomor sudah di generate"
                    })
                })
            
        try:
            latest_item_number = Item.objects.filter(bundle__box__yeardate=bundle.box.yeardate).latest('item_number').item_number + 1
        except:
            latest_item_number = 1
        form = ItemForm(initial={'yeardate': bundle.yeardate, 'item_number': latest_item_number, 'bundle':  Bundle.objects.first().id})
    return render(request, 'arsip_tata/item_form.html', {
        'form': form,
    })

def edit_item(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == "POST":
        # form = CustomerForm(request.POST, request.FILES, instance=customer)
        form = ItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            itemsave = form.save(commit=False)
            itemsave.total = itemsave.copy + itemsave.original
            if item.bundle.box.isgen:
                item.issync = False
            # bundle = Bundle.objects.get(id=item.bundle_id)
            # itemsave.codegen = "-".join([str(itemsave.yeardate), str(item.bundle.box.box_number), str(item.bundle.bundle_number), str(itemsave.item_number)])
            itemsave.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "itemListChanged": None,
                        "showMessage": f"{itemsave.item_number} updated."
                    })
                }
            )
    else:
        form = ItemForm(instance=item)
    return render(request, 'arsip_tata/item_form.html', {
        'form': form,
        'item': item,
        'module': 'Edit Data'
    })

@require_http_methods(['DELETE'])
def remove_item(request, pk):
    item = get_object_or_404(Item, pk=pk)
    item.delete()
    return HttpResponse(
        status=204,
        headers={
            'HX-Trigger': json.dumps({
                "itemListChanged": None,
                "showMessage": f"{item.item_number} deleted."
            })
        })

def create_xls(datalist, sheet, year):
    # wb = Workbook()
    # sheet = wb.active
    sheet.title = str(year)
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['B'].width = 7
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 60
    sheet.column_dimensions['F'].width = 7
    sheet.column_dimensions['G'].width = 2
    sheet.column_dimensions['H'].width = 6 
    sheet.column_dimensions['I'].width = 7 
    sheet.column_dimensions['J'].width = 7
    sheet.column_dimensions['K'].width = 7
    sheet.column_dimensions['L'].width = 30

    sheet.merge_cells('A1:L1')
    sheet.merge_cells('A2:L2')
    sheet.merge_cells('A3:L3')
    aligncenter = Alignment(horizontal='center')
    headerfont = Font(name='Arial Narrow', size=14, bold=True)
    sheet['A1'] = "DAFTAR ARSIP INAKTIF"
    sheet['A1'].alignment = aligncenter
    sheet['A1'].font = headerfont
    sheet['A2'] = "UNIT PENGOLAH: BALAI WILAYAH SUNGAI MALUKU UTARA"
    sheet['A2'].alignment = aligncenter
    sheet['A2'].font = headerfont
    sheet['A3'] = "TAHUN PENATAAN {}".format(year)
    sheet['A3'].alignment = aligncenter
    sheet['A3'].font = headerfont
    
    centervh = Alignment(horizontal='center', vertical='center', wrap_text=True)
    centerv = Alignment(vertical='center', wrap_text=True)
    wraptxt = Alignment(wrap_text=True)

    thin_border1 = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    thin_border2 = Border(bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
    thin_border3 = Border(top=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
    thin_border4 = Border(right=Side(style='thin'),left=Side(style='thin'))
    thin_border5 = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    thin_border6 = Border(top=Side(style='thin'))
    thin_border7 = Border(bottom=Side(style='thin'))
    thin_border8 = Border(left=Side(style='thin'))
    thin_border9 = Border(right=Side(style='thin'))


    font_style1 = Font(name='Arial Narrow', size=11, bold=True)
    font_style2 = Font(name='Arial', size=8.5)
    font_style3 = Font(name='Arial', size=8.5, italic=True)
    color1 = PatternFill(start_color="c6d5f7", fill_type = "solid")

    # sheet.row_dimensions[7].height = 28
    for cell in sheet["7:7"]:
        cell.alignment = centervh
        cell.font = font_style1
    for cell in sheet["8:8"]:
        cell.alignment = centervh
        cell.font = font_style1

    sheet.merge_cells('A7:A8')
    sheet.merge_cells('B7:B8')
    sheet.merge_cells('C7:C8')
    sheet.merge_cells('D7:D8')
    sheet.merge_cells('E7:E8')
    sheet.merge_cells('F7:F8')
    sheet.merge_cells('G7:H8')
    sheet.merge_cells('I7:K7')
    sheet.merge_cells('L7:L8')
    sheet.merge_cells('I8:J8')
    sheet.merge_cells('G9:H9')
    sheet.merge_cells('I9:J9')

    
    headers = ("NO BRKS", "NO URUT", "KODE", "INDEKS/JUDUL", "URAIAN MASALAH / KEGIATAN", "TAHUN")
    for i in headers:
        sheet.cell(row=7, column=headers.index(i)+1).value = i
        sheet.cell(row=7, column=headers.index(i)+1).border = thin_border1
        sheet.cell(row=8, column=headers.index(i)+1).border = thin_border1
    no = 0
    for i in range(1, 13):
        if i not in [8,10]:
            no += 1
            sheet.cell(row=9, column=i).value = no
            
        sheet.cell(row=9, column=i).alignment = centervh
        sheet.cell(row=9, column=i).border = thin_border1
        sheet.cell(row=9, column=i).font = Font(name='Arial Narrow', size=8, bold=True)
        sheet.cell(row=7, column=i).fill = color1
        sheet.cell(row=8, column=i).fill = color1
        sheet.cell(row=9, column=i).fill = color1

    sheet.row_dimensions[9].height = 9
    
    sheet.cell(row=7, column=7).value = "JUMLAH BERKAS"
    sheet.cell(row=7, column=7).border = thin_border1
    sheet.cell(row=8, column=7).border = thin_border1
    sheet.cell(row=7, column=8).border = thin_border1
    sheet.cell(row=8, column=8).border = thin_border1

    sheet.cell(row=7, column=9).value = "KETERANGAN"
    sheet.cell(row=7, column=9).border = thin_border1
    sheet.cell(row=7, column=10).border = thin_border1
    sheet.cell(row=7, column=11).border = thin_border1


    sheet.cell(row=8, column=9).value = "ASLI/COPY"
    sheet.cell(row=8, column=9).border = thin_border1
    sheet.cell(row=8, column=10).border = thin_border1
    
    sheet.cell(row=8, column=11).value = "BOX"
    
    sheet.cell(row=8, column=9).border = thin_border1

    sheet.cell(row=8, column=10).border = thin_border1
    sheet.cell(row=8, column=11).border = thin_border1
    
    
    sheet.cell(row=8, column=9).font = font_style1
    sheet.cell(row=8, column=10).font = font_style1
    sheet.cell(row=8, column=11).font = font_style1
    sheet.cell(row=8, column=12).font = font_style1
    
    sheet.cell(row=7, column=12).value = "KLASIFIKASI KEAMANAN DAN AKSES ARSIP DINAMIS"
    sheet.cell(row=7, column=12).border = thin_border1
    sheet.cell(row=8, column=12).border = thin_border1
    sheet.cell(row=7, column=12).alignment = centervh

    i = 10
    for res in datalist:
        sheet['{}{}'.format('K', i)].border = thin_border4
        sheet['{}{}'.format('A', i)].border = thin_border4
        sheet['{}{}'.format('B', i)].border = thin_border4
        sheet['{}{}'.format('C', i)].border = thin_border4
        sheet['{}{}'.format('D', i)].border = thin_border4
        sheet['{}{}'.format('E', i)].border = thin_border4
        sheet['{}{}'.format('F', i)].border = thin_border4
        sheet['{}{}'.format('G', i)].border = thin_border7
        sheet['{}{}'.format('H', i)].border = thin_border7
        sheet['{}{}'.format('I', i)].border = thin_border4
        sheet['{}{}'.format('J', i)].border = thin_border4
        sheet['{}{}'.format('L', i)].border = thin_border4

        sheet['{}{}'.format('B', i)].border = thin_border1
        sheet['{}{}'.format('E', i)].border = thin_border1
        # sheet['{}{}'.format('G', i)].border = thin_border7
        # sheet['{}{}'.format('H', i)].border = thin_border1
        sheet['{}{}'.format('I', i)].border = thin_border1
        sheet['{}{}'.format('J', i)].border = thin_border1
        sheet['{}{}'.format('L', i)].border = thin_border1
            
        sheet['{}{}'.format('A', i)].value = res[0]
        sheet['{}{}'.format('B', i)].value = res[1]
        sheet['{}{}'.format('C', i)].value = res[2]
        sheet['{}{}'.format('D', i)].value = res[3]
        sheet['{}{}'.format('E', i)].value = res[4]
        sheet['{}{}'.format('F', i)].value = res[5]
        sheet['{}{}'.format('H', i)].value = res[6]
        sheet['{}{}'.format('I', i)].value = res[7]
        sheet['{}{}'.format('J', i)].value = res[8]
        sheet['{}{}'.format('K', i)].value = res[9]
        sheet['{}{}'.format('L', i)].value = res[10]
        sheet['{}{}'.format('K', i)].alignment = centervh
        sheet['{}{}'.format('A', i)].alignment = centervh
        sheet['{}{}'.format('B', i)].alignment = centervh
        sheet['{}{}'.format('C', i)].alignment = centervh
        sheet['{}{}'.format('D', i)].alignment = centervh
        sheet['{}{}'.format('E', i)].alignment = centerv
        sheet['{}{}'.format('F', i)].alignment = centervh
        # sheet['{}{}'.format('G', i)].alignment = centervh
        sheet['{}{}'.format('H', i)].alignment = centervh
        sheet['{}{}'.format('I', i)].alignment = centervh
        sheet['{}{}'.format('J', i)].alignment = centervh
        sheet['{}{}'.format('L', i)].alignment = centervh
        if res[0] != '':
            sheet['{}{}'.format('A', i)].border = thin_border3
            sheet['{}{}'.format('C', i)].border = thin_border3
            sheet['{}{}'.format('D', i)].border = thin_border3
            sheet['{}{}'.format('F', i)].border = thin_border3
        if res[9] != '':
            sheet['{}{}'.format('K', i)].border = thin_border3

        i += 1

    sheet['{}{}'.format('A', i)].border = thin_border6
    sheet['{}{}'.format('C', i)].border = thin_border6
    sheet['{}{}'.format('D', i)].border = thin_border6
    sheet['{}{}'.format('F', i)].border = thin_border6
    sheet['{}{}'.format('K', i)].border = thin_border6

def generate_data(year):
    result = []
    yeardata = Year.objects.get(yeardate=year)
    for box in yeardata.box_set.all():
        cbox = True
        for bundle in box.bundle_set.all():
            cbundle = True
            for item in bundle.item_set.all().order_by('item_number'):
                if cbox:
                    boxnumber = box.box_number
                    cbox = False
                else:
                    boxnumber = ''

                if cbundle:
                    bundlenumber = bundle.bundle_number
                    code = bundle.code
                    creator = bundle.creator
                    description = f"{item.title}\n{bundle.description}"
                    year_bundle = bundle.year_bundle
                    cbundle = False
                else:
                    bundlenumber = ''
                    code = ''
                    creator = ''
                    description = item.title
                    year_bundle = ''
                
                dataset = (bundlenumber, item.item_number, code, creator, description, year_bundle, item.total, item.original, item.copy, boxnumber, item.get_accesstype_display())
                result.append(dataset)
    return result

def report(request, year):
    res = generate_data(year)
    wb = Workbook()
    filename = f"data_{__package__.split('.')[1]}_tahun_penataan_{year}.xlsx"
    sheet = wb.active
    create_xls(datalist=res, sheet=sheet, year=year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)    
    wb.save(response)
    return response

def generate_data_perbox(year, box_number):
    result = []
    yeardata = Year.objects.get(yeardate=year)
    for box in yeardata.box_set.filter(box_number=box_number):
        cbox = True
        for bundle in box.bundle_set.all():
            cbundle = True
            for item in bundle.item_set.all().order_by('item_number'):
                if cbox:
                    boxnumber = box.box_number
                    cbox = False
                else:
                    boxnumber = ''

                if cbundle:
                    bundlenumber = bundle.bundle_number
                    code = bundle.code
                    creator = bundle.creator
                    description = f"{item.title}\n{bundle.description}"
                    year_bundle = bundle.year_bundle
                    cbundle = False
                else:
                    bundlenumber = ''
                    code = ''
                    creator = ''
                    description = item.title
                    year_bundle = ''
                
                dataset = (bundlenumber, item.item_number, code, creator, description, year_bundle, item.total, item.original, item.copy, boxnumber, item.get_accesstype_display())
                result.append(dataset)
    return result

def report_perbox(request, year, box_number):
    res = generate_data_perbox(year, box_number)
    wb = Workbook()
    filename = f"data_{__package__.split('.')[1]}_tahun_penataan_{year}_box_{box_number}.xlsx"
    sheet = wb.active
    create_xls(datalist=res, sheet=sheet, year=year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)    
    wb.save(response)
    return response

def label_perbox(request, year, box_number):
    if not request.user.is_authenticated:
        return redirect('login')
    box = Box.objects.get(box_number=box_number, yeardate=year)
   
    bundles = Bundle.objects.filter(box_id=box.id).all()
    bundle_numbers = []
    yearbundle = ""
    code = ""
    bundle_years = []
    codes = []
    if len(bundles) != 0:
        yearbundle = str(bundles[0].year_bundle)
        code = str(bundles[0].code)
    for bundle in bundles:
        bundle_numbers.append(str(bundle.bundle_number))
        bundle_years.append(str(bundle.year_bundle))
        codes.append(str(bundle.code))
        items = Item.objects.filter(bundle_id=bundle.id).all()
        item_numbers = []
        for item in items:
            item_numbers.append(item.item_number)
    bundle_numbers.sort()
    item_numbers.sort()
    bundle_years.sort()
    minitem = "__"
    maxitem = "__"
    if len(item_numbers) != 0:
        minitem = str(item_numbers[0])
        maxitem = str(item_numbers[-1])
    pdf = io.BytesIO()
    doc = SimpleDocTemplate(pdf, pagesize=A5)

    frame = Frame(doc.leftMargin-60, doc.bottomMargin, doc.width, doc.height)
    template = PageTemplate(frames=[frame], id='mytemplate')

    doc.addPageTemplates([template])
    elements = []
    mydata = []
    # mydata.append(("No", "Kode", "Judul Dokumen"))
    c_width = [2.3*inch, 0.2*inch, 2*inch]
    # c_width = [3*inch, 0.2*inch, 2*inch]

    
    stylesample = getSampleStyleSheet()
    style2 = stylesample["Heading1"]
    style2.wordWrap = 'CJK'
    style3 = stylesample["Heading2"]
    style3.wordWrap = 'CJK'

    filename = f"boxlabel_{year}_{box_number}.pdf"
    myset = (Paragraph("Tahun Penataan {}".format(year), stylesample["Italic"]), Paragraph("", style2), QRCodeImage(f"{settings.DOMAIN}/arsip_tata/search_qrcode/{year}/{box_number}", size=25 * mm), )
    mydata.append(myset)
    myset = (Paragraph("NO. BOX", style2), Paragraph(":", style2), Paragraph(str(box_number), style2))
    mydata.append(myset)
    myset = (Paragraph("NO. BERKAS", style2), Paragraph(":", style2), Paragraph(", ".join(bundle_numbers), style2))
    mydata.append(myset)
    myset = (Paragraph("KODE", style2), Paragraph(":", style2), Paragraph(", ".join(list(set(codes))), style3))
    mydata.append(myset)
    myset = (Paragraph("NO. ITEM", style2), Paragraph(":", style2), Paragraph(f"{minitem} - {maxitem}", style2))
    mydata.append(myset)
    myset = (Paragraph("TAHUN", style2), Paragraph(":", style2), Paragraph(", ".join(list(set(bundle_years))), style2))
    mydata.append(myset)
    
    mytable = Table(mydata, colWidths=c_width, hAlign='LEFT')
    mytable.setStyle(TableStyle([
                       ('FONTSIZE',(0,0),(-1,0),16),
                       ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                       ('VALIGN',(0, 0),(-1,-1),'TOP'),
                       ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                       ('TOPPADDING', (0, 0), (-1, -1), 7),
                       ]))
    # elements.append(Spacer(1, 5))
    elements.append(mytable)


    doc.build(elements)
    pdf.seek(0)
    response = HttpResponse(pdf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline;filename={filename}'
    # response['Content-Disposition'] = f'attachment;filename={filename}'

    return response

def label_perbundle(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')
    bundle = Bundle.objects.get(pk=pk)
    pdf = io.BytesIO()
    doc = SimpleDocTemplate(pdf, pagesize=A5)

    frame = Frame(doc.leftMargin-50, doc.bottomMargin+30, doc.width, doc.height)
    template = PageTemplate(frames=[frame], id='mytemplate')

    doc.addPageTemplates([template])
    elements = []
    mydata = []
    # mydata.append(("No", "Kode", "Judul Dokumen"))
    c_width = [1.1*inch, 0.2*inch, 3.5*inch]
    
    stylesample = getSampleStyleSheet()
    style2 = stylesample["Italic"]
    style2.wordWrap = 'CJK'
    filename = f"bundlelabel_{bundle.yeardate}_{bundle.bundle_number}.pdf"
    myset = (Paragraph("NO. BOX", style2), Paragraph(":", style2), Paragraph(str(bundle.box.box_number), style2))
    mydata.append(myset)
    myset = (Paragraph("NO. BERKAS", style2), Paragraph(":", style2), Paragraph(str(bundle.bundle_number), style2))
    mydata.append(myset)
    myset = (Paragraph("PEKERJAAN", style2), Paragraph(":", style2), Paragraph(str(bundle.description), style2))
    mydata.append(myset)
    myset = (Paragraph("KODE", style2), Paragraph(":", style2), Paragraph(str(bundle.code), style2))
    mydata.append(myset)
    myset = (Paragraph("INDEX", style2), Paragraph(":", style2), Paragraph(str(bundle.creator), style2))
    mydata.append(myset)
    myset = (Paragraph("TAHUN", style2), Paragraph(":", style2), Paragraph(str(bundle.year_bundle), style2))
    mydata.append(myset)
    
    mytable = Table(mydata, colWidths=c_width, hAlign='LEFT')
    mytable.setStyle(TableStyle([
                    #    ('FONTSIZE',(0,0),(-1,0),16),
                       ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                       ('VALIGN',(0, 0),(-1,-1),'TOP'),
                       ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                       ('TOPPADDING', (0, 0), (-1, -1), 3),
                       ]))
    # elements.append(Spacer(1, 5))
    elements.append(mytable)
    mydata = []
    c_width = [0.5*inch, 3.8*inch, 0.5*inch]
    style2 = stylesample["Normal"]
    style2.wordWrap = 'CJK'

    myset = (Paragraph("ITEM", style2), Paragraph("URAIAN MASALAH KEGIATAN", style2), Paragraph("JML", style2))
    mydata.append(myset)

    for item in bundle.item_set.all():
        myset = (Paragraph(str(item.item_number), style2), Paragraph(str(item.title), style2), Paragraph(str(item.total), style2))
        mydata.append(myset)
    mytable = Table(mydata, colWidths=c_width, hAlign='LEFT')
    mytable.setStyle(TableStyle([
                       ('FONTSIZE',(0,0),(-1,0),16),
                       ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                       ('VALIGN',(0, 0),(-1,-1),'TOP'),
                       ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                       ('TOPPADDING', (0, 0), (-1, -1), 2),
                       ]))
    elements.append(mytable)
    doc.build(elements)
    pdf.seek(0)
    response = HttpResponse(pdf.read(), content_type='application/pdf', )
    response['Content-Disposition'] = f'inline;filename={filename}'
    return response

def search_qrcode(request, year, box_number):
    if not request.user.is_authenticated:
        res = Box.objects.get(box_number=box_number, yeardate=year)
        return render(request=request, template_name='arsip_tata/search_qrcode.html', context={'datalist': res})
    else:
        # TODO: ADD MORE FEATURE FOR LOGIN USER
        res = Box.objects.get(box_number=box_number, yeardate=year)
        return render(request=request, template_name='arsip_tata/search_qrcode.html', context={'datalist': res})
        
def show_customers(request):
    if not request.user.is_authenticated:
        return redirect('login')
    context = {
    }
    return render(request=request, template_name='arsip_tata/show_customer.html', context=context)

def customer_list(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'arsip_tata/customer_list.html', {
        'customers': Customer.objects.all(),
    })

def add_customer(request):
    if request.method == "POST":
        form = CustomerForm(request.POST or None, request.FILES or None)
        # if request.is_ajax():
        # print(form)
        if form.is_valid():
            customer = form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "customerListChanged": None,
                        "showMessage": f"{customer.name} added."
                    })
                })
    else:
        form = CustomerForm()
    return render(request, 'arsip_tata/customer_form.html', {
        'form': form,
        'module': 'Tambah Data'
    })

def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == "POST":
        form = CustomerForm(request.POST, request.FILES, instance=customer)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "customerListChanged": None,
                        "showMessage": f"{customer.name} updated."
                    })
                }
            )
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'arsip_tata/customer_form.html', {
        'form': form,
        'customer': customer,
        'module': 'Edit Data'
    })

@require_http_methods(['DELETE'])
def remove_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer.photo.delete(save=True)
    customer.idcard.delete(save=True)
    customer.delete()
    return HttpResponse(
        status=204,
        headers={
            'HX-Trigger': json.dumps({
                "customerListChanged": None,
                "showMessage": f"{customer.name} deleted."
            })
        })

def show_trans(request):
    if not request.user.is_authenticated:
        return redirect('login')
    context = {
    }
    return render(request=request, template_name='arsip_tata/show_trans.html', context=context)

def trans_list(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'arsip_tata/trans_list.html', {
        'trans': Trans.objects.all().order_by("-id"),
    })

def add_trans(request):
    if request.method == "POST":
        form = TransForm(request.POST)
        if form.is_valid():
            trans = form.save(commit=False)
            yearnow = datetime.today().year
            stryearnow = str(yearnow)
            ctran = Trans.objects.filter(codetrans__contains=stryearnow).order_by("-id").first()
            if not ctran:
                no = 1
            else:
                no = int(ctran.codetrans[4:]) + 1
            print(stryearnow + "{:05d}".format(no))
            trans.codetrans = stryearnow + "{:04d}".format(no)
            trans.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "transListChanged": None,
                        "showMessage": f"{trans.id} added."
                    })
                })
    else:
        form = TransForm()
    return render(request, 'arsip_tata/trans_form.html', {
        'form': form,
        'module': 'Tambah Data'
    })

def edit_trans(request, pk):
    trans = get_object_or_404(Trans, pk=pk)
    if request.method == "POST":
        form = TransForm(request.POST, instance=trans)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "transListChanged": None,
                        "showMessage": f"{trans.id} updated."
                    })
                }
            )
    else:
        form = TransForm(instance=trans)
    return render(request, 'arsip_tata/trans_form.html', {
        'form': form,
        'trans': trans,
        'module': 'Edit Data'
    })

@ require_POST
def remove_trans(request, pk):
    trans = get_object_or_404(Trans, pk=pk)
    trans.delete()
    return HttpResponse(
        status=204,
        headers={
            'HX-Trigger': json.dumps({
                "transListChanged": None,
                "showMessage": f"{trans.id} deleted."
            })
        })

def show_trans_detail(request, trans_id):
    if not request.user.is_authenticated:
        return redirect('login')
    trans = Trans.objects.get(id=trans_id)
    context = {
        'trans': trans,
    }
    return render(request=request, template_name='arsip_tata/show_trans_detail.html', context=context)

@csrf_exempt
def trans_detail_list(request, trans_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    transdetail = TransDetail.objects.filter(trans_id=trans_id)
    return render(request, 'arsip_tata/trans_detail_list.html', {
        'transdetail': transdetail,
    })

@csrf_exempt
def add_trans_detail_old(request, trans_id):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == "POST":
        code = request.POST.get('code')
        # form = AddTransDetailForm(request.POST)
        try:
            # year, box, bundle, item = code.split("-")
            message = f"Kode item berkas {code} sukses"
            item = Item.objects.get(codegen=code)
            if item:
                if item.total == 1:
                    message = f"Kode item berkas {code} hanya ada 1, tidak boleh pinjam lebih dari 1 hari"
                
                td = TransDetail.objects.filter(item_id=item.id, date_return__isnull=True)
                if td.count() != 0:
                    message = f"Kode item berkas {code} sedang dipinjam"
                else:
                    td = TransDetail(item_id=item.id, trans_id=trans_id)
                    td.save()
                    message = f"Kode item berkas {code} tersimpan"
        except:
            message = f"Kode item berkas {code} tidak ada"
        return HttpResponse(
            status=204,
            headers={
                'HX-Trigger': json.dumps({
                    "transDetailListChanged": None,
                    "showMessage": message
                })
            })
    else:
        form = AddTransDetailForm()
    return render(request, 'arsip_tata/add_trans_detail_form.html', {
        'form': form,
        'module': 'Tambah Data'
    })

@csrf_exempt
def add_trans_detail(request, trans_id):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == "POST":
        code = request.POST.get('code')
        # form = AddTransDetailForm(request.POST)
        try:
            # year, box, bundle, item = code.split("-")
            message = f"Kode item berkas {code} sukses"
            item = Item.objects.get(codegen=code)
            if item:
                if item.total == 1:
                    message = f"Kode item berkas {code} hanya ada 1, tidak boleh pinjam lebih dari 1 hari"
                
                td = TransDetail.objects.filter(item_id=item.id, date_return__isnull=True)
                if td.count() != 0:
                    message = f"Kode item berkas {code} sedang dipinjam"
                else:
                    td = TransDetail(item_id=item.id, trans_id=trans_id)
                    td.save()
                    message = f"Kode item berkas {code} tersimpan"
        except:
            message = f"Kode item berkas {code} tidak ada"
        return HttpResponse(
            status=204,
            headers={
                'HX-Trigger': json.dumps({
                    "transListChanged": None,
                    "showMessage": message
                })
            })
    else:
        form = AddTransDetailForm()
    return render(request, 'arsip_tata/add_trans_detail_form.html', {
        'form': form,
        'module': 'Tambah Data'
    })

@csrf_exempt
def customer_info(request, customer_id):
    if not request.user.is_authenticated:
        return redirect('login')
    customer = Customer.objects.get(id=customer_id)

    return render(request, 'arsip_tata/customer_info.html', {
        'data': customer,
    })

@csrf_exempt
def item_info(request, item_id):
    if not request.user.is_authenticated:
        return redirect('login')
    item = Item.objects.get(id=item_id)

    return render(request, 'arsip_tata/item_info.html', {
        'data': item,
    })

# @ require_POST
def remove_transdetail(request, pk):
    if request.method == "POST":
        trans = get_object_or_404(TransDetail, pk=pk)
        trans_id = trans.trans.id
        trans.delete()
        # return redirect('arsip_tata_show_trans_detail', trans_id=trans_id)
        return redirect('arsip_tata_show_trans')

def trans_form(request, pk):
    trans = Trans.objects.get(pk=pk)
    detail = trans.transdetail_set.all()
    pdf = io.BytesIO()
    doc = SimpleDocTemplate(pdf, pagesize=A4)
    date1 = trans.date_trans
    date1show = date1.strftime('%d %B %Y')
    date2 = date1 + timedelta(days=7)
    date2show = date2.strftime('%d %B %Y')
    styles = getSampleStyleSheet()
    title = "Form Peminjaman Dokumen"

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height)
    template = PageTemplate(frames=[frame], id='mytemplate')

    doc.addPageTemplates([template])
    elements = []
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f'Kode Peminjaman: <strong>{trans.codetrans}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Nama Peminjam: <strong>{trans.customer.name}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Satuan Kerja: <strong>{trans.customer.organization}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'No WhatsApp: <strong>{trans.customer.phone_number}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Tanggal Pinjam: <strong>{date1show}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Akan Kembali pada: <strong>{date2show}</strong>', styles['Normal']))
    mydata = []
    mydata.append(("No", "Kode", "Judul Dokumen"))
    c_width = [0.4*inch, 1*inch, 5*inch]
    
    style2 = getSampleStyleSheet()
    style2 = style2["BodyText"]
    style2.wordWrap = 'CJK'
    filename = f"form_pinjam_{trans.codetrans}.pdf"
    for idx, data in enumerate(detail):
        myset = (Paragraph(str(idx+1), style2) , Paragraph(data.item.codegen, style2), Paragraph(data.item.title + "<br/>" + data.item.bundle.description.replace("\n", "<br/>") , style2))
        mydata.append(myset)
    
    mytable = Table(mydata, colWidths=c_width, hAlign='LEFT')
    mytable.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightgreen),
                       ('INNERGRID', (0,0), (-1,-1), 0.25, colors.red),
                       ('FONTSIZE',(0,0),(-1,0),12),
                       ('FONTSIZE',(0,1),(-1,-1),8),
                       ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                       ('ALIGN', (0,0), (-1,0), 'CENTER'),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                       ('VALIGN',(0,0),(-1,-1),'TOP'),
                       ]))
    elements.append(Spacer(1, 10))
    elements.append(mytable)

    elements.append(Spacer(1, 32))
    today = datetime.today().strftime('%d %B %Y')
    rightal = ParagraphStyle(name="RightAl",alignment=TA_RIGHT)
    elements.append(Paragraph(f'Ternate, {today}', rightal))
    elements.append(Spacer(1, 32))
    elements.append(Paragraph('(Petugas Arsip)&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;', rightal))

    doc.build(elements)
    pdf.seek(0)
    response = HttpResponse(pdf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline;filename={filename}'
    return response

def show_transret(request):
    if not request.user.is_authenticated:
        return redirect('login')
    context = {
    }
    return render(request=request, template_name='arsip_tata/show_transret.html', context=context)

def transret_list(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'arsip_tata/transret_list.html', {
        'trans': Trans.objects.all().order_by("-id"),
    })

def show_transret_detail(request, trans_id):
    if not request.user.is_authenticated:
        return redirect('login')
    trans = Trans.objects.get(id=trans_id)
    context = {
        'trans': trans,
    }
    return render(request=request, template_name='arsip_tata/show_transret_detail.html', context=context)

def edit_transdetail(request, pk):
    trans = get_object_or_404(TransDetail, pk=pk)
    if request.method == "POST":
        form = EditTransDetailForm(request.POST, instance=trans)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "transListChanged": None,
                        "showMessage": f"{trans.id} updated."
                    })
                }
            )
    else:
        if trans.date_return == None:
            form = EditTransDetailForm(instance=trans)
        else:
            trans.date_return = None
            trans.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "transListChanged": None,
                        "showMessage": f"{trans.id} updated."
                    })
                }
            )

    return render(request, 'arsip_tata/edit_trans_detail_form.html', {
        'form': form,
        'transdetail': trans,
        'module': 'Edit Data'
    })


def edit_transdetail_old(request, pk):
    trans = get_object_or_404(TransDetail, pk=pk)
    if request.method == "POST":
        form = EditTransDetailForm(request.POST, instance=trans)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "transretDetailListChanged": None,
                        "showMessage": f"{trans.id} updated."
                    })
                }
            )
    else:
        if trans.date_return == None:
            form = EditTransDetailForm(instance=trans)
        else:
            trans.date_return = None
            trans.save()
            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "transretDetailListChanged": None,
                        "showMessage": f"{trans.id} updated."
                    })
                }
            )

    return render(request, 'arsip_tata/edit_trans_detail_form.html', {
        'form': form,
        'transdetail': trans,
        'module': 'Edit Data'
    })

@csrf_exempt
def transret_detail_list(request, trans_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    transdetail = TransDetail.objects.filter(trans_id=trans_id)
    return render(request, 'arsip_tata/transret_detail_list.html', {
        'transdetail': transdetail,
    })

def transret_form(request, pk):
    trans = Trans.objects.get(pk=pk)
    detail = trans.transdetail_set.all()
    pdf = io.BytesIO()
    doc = SimpleDocTemplate(pdf, pagesize=A4)
    date1 = trans.date_trans
    date1show = date1.strftime('%d %B %Y')
    styles = getSampleStyleSheet()
    title = "Form Pengembalian Dokumen"

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height)
    template = PageTemplate(frames=[frame], id='mytemplate')

    doc.addPageTemplates([template])
    elements = []
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f'Kode Peminjaman: <strong>{trans.codetrans}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Nama Peminjam: <strong>{trans.customer.name}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Satuan Kerja: <strong>{trans.customer.organization}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'No WhatsApp: <strong>{trans.customer.phone_number}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'Tanggal Pinjam: <strong>{date1show}</strong>', styles['Normal']))
    elements.append(Spacer(1, 6))
    mydata = []
    mydata.append(("No", "Kode", "Judul Dokumen", "Kembali"))
    c_width = [0.4*inch, 1*inch, 4*inch, 1*inch]
    
    style2 = getSampleStyleSheet()
    style2 = style2["BodyText"]
    style2.wordWrap = 'CJK'
    filename = f"form_kembali_{trans.codetrans}.pdf"
    for idx, data in enumerate(detail):
        if data.date_return:
            date_return = data.date_return.strftime('%d %b %Y')
        # else:
        #     date_return = ""
            myset = (Paragraph(str(idx+1), style2) , Paragraph(data.item.codegen, style2), Paragraph(data.item.title + "<br/>" + data.item.bundle.description.replace("\n", "<br/>") , style2), Paragraph(date_return, style2))
            mydata.append(myset)

    mytable = Table(mydata, colWidths=c_width, hAlign='LEFT')
    mytable.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightblue),
                       ('INNERGRID', (0,0), (-1,-1), 0.25, colors.red),
                       ('FONTSIZE',(0,0),(-1,0),12),
                       ('FONTSIZE',(0,1),(-1,-1),8),
                       ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                       ('ALIGN', (0,0), (-1,0), 'CENTER'),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                       ('VALIGN',(0,0),(-1,-1),'TOP'),
                       ]))
    elements.append(Spacer(1, 10))
    elements.append(mytable)
    elements.append(Spacer(1, 32))
    today = datetime.today().strftime('%d %B %Y')
    rightal = ParagraphStyle(name="RightAl",alignment=TA_RIGHT)
    elements.append(Paragraph(f'Ternate, {today}', rightal))
    elements.append(Spacer(1, 32))
    elements.append(Paragraph('(Petugas Arsip)&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;', rightal))

    doc.build(elements)
    pdf.seek(0)
    response = HttpResponse(pdf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline;filename={filename}'
    return response

def search_item(request):
    context = {}
    if request.GET.get("title") or request.GET.get("description"):
        title = request.GET.get("title")
        description = request.GET.get("description")
        context['description'] = description 
        context['title'] = title 

        # print(description)
        if title != None and description == None:
            items = Item.objects.filter(title__icontains=title)
        elif title == None and description != None:
            items = Item.objects.filter(bundle__description__icontains=description)
        else:
            items = Item.objects.filter(Q(bundle__description__icontains=description) & Q(title__icontains=title))       
    else:
        items = Item.objects.all()

    if request.GET.get("page"):
        page = request.GET.get("page")
    else:
        page = 1
    paginator = Paginator(items, 100)
    
    try:
        items = paginator.page(page)
    except PageNotAnInteger:
        items = paginator.page(1)
    except EmptyPage:
        items = paginator.page(paginator.num_pages)
    context['data'] = {}
    data = []
    for item in items:
        status = 'Ada'
        trans_id = None
        if item.transdetail_set.filter(date_return__isnull=True).count() != 0:
            status='Dipinjam'
            trans_id = item.transdetail_set.filter(date_return__isnull=True).first().trans_id
        myset = (item.codegen, item.title, item.bundle.description, status, trans_id)
        data.append(myset)
    context['data'] = data
    context['has_other_pages'] = items.has_other_pages()
    try:
        context['has_previous'] = items.has_previous()
        context['previous_page_number'] = items.previous_page_number()
    except:
        context['previous_page_number'] = False
    
    context['number'] = items.number
    context['page_range'] = items.paginator.page_range
    try:
        context['has_next'] = items.has_next()
        context['next_page_number'] = items.next_page_number()
    except:
        context['next_page_number'] = False
    context['num_pages'] = items.paginator.num_pages    

    context['form'] = SearchItemForm()
    return render(request,'arsip_tata/search_item_form.html', context=context)

@csrf_exempt
def item_upload_pdf(request):
    if request.method == "POST":
        if request.FILES:
            item_id = request.POST.get("item_id")
            item = Item.objects.get(id=item_id)
            filename = f"{__package__.split('.')[1]}$${item.codegen}.pdf"
            upload = request.FILES.getlist('uploadfile')[0]
            tmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", filename)
            fss = FileSystemStorage()
            if exists(tmppath):
                os.remove(tmppath)
            fss.save(tmppath, upload)
            item = Item.objects.get(id=item_id)
            item.uploaded_by = request.user
            item.uploaded_date = datetime.now()
            item.save()

            return HttpResponse(
                status=204,
                headers={
                    'HX-Trigger': json.dumps({
                        "movieListChanged": None,
                        "showMessage": 'Upload File Sukses, tunggu beberapa saat kemudian refresh halaman'
                    })
                })
    else:
        item_id = request.GET.get("item_id")
        # year = str(request.GET.get("year"))
        # folder = request.GET.get("folder")
        # print(item_id)
    return render(request, 'arsip_tata/item_upload_pdf.html', {
        'item_id': item_id,
    })

@csrf_exempt
@user_passes_test(lambda user: Group.objects.get(name='admin') in user.groups.all())
def box_sync(request, pk):
    if request.method == "POST":
        box = get_object_or_404(Box, pk=pk)
        try:
            prevbox = Box.objects.get(box_number=str(int(box.box_number)-1), yeardate=box.yeardate)
            # if prevbox:
            bundles = Bundle.objects.filter(box=prevbox)
            bundle_maxnumber = bundles.aggregate(Max("bundle_number"))['bundle_number__max']
            bundle_maxid = Bundle.objects.filter(bundle_number=bundle_maxnumber, yeardate=prevbox.yeardate).first().id
            item_maxnumber = Item.objects.filter(bundle_id=bundle_maxid).aggregate(Max("item_number"))['item_number__max']
            bundle_number = bundle_maxnumber+1
            item_number = item_maxnumber+1
        except:
            bundle_number = 1
            item_number = 1
        bundles = Bundle.objects.filter(box=box)
        # cek isi bundle harus ada isi semua
        errorfound = False
        for bundle in bundles:
            if Item.objects.filter(bundle=bundle).count() == 0:
                errorfound = True
        if not errorfound:
            itemcounter = 0
            for idx, bundle in enumerate(bundles):
                bundle.bundle_number = bundle_number+idx
                bundle.syncstatus = 2
                bundle.save()
                items = Item.objects.filter(bundle=bundle)
                for item in items:
                    item.item_number = item_number+itemcounter
                    itemcounter += 1
                    item.save()
            # box.token = 'PROSES'
            box.isgen = True
            box.save()

            message = "Nomor Box, Berkas dan Item berhasil di generate"
        else:
            message = "Generate nomor Gagal, Ada berkas yang kosong"
        
        return HttpResponse(
            status=204,
            headers={
                'HX-Trigger': json.dumps({
                    "boxListChanged": None,
                    "showMessage": message
                })
            })
        
    return render(request, 'arsip_tata/box_sync.html', {
    })

@csrf_exempt
@user_passes_test(lambda user: Group.objects.get(name='admin') in user.groups.all())
def bundle_sync(request, pk):

    if request.method == "POST":
        bundle = get_object_or_404(Bundle, pk=pk)
        bundle.syncstatus = 2
        bundle.save()

        return HttpResponse(
            status=204,
            headers={
                'HX-Trigger': json.dumps({
                    "bundleListChanged": None,
                    "showMessage": "Sinkronisasi data sedang di proses..."
                })
            })
        
    return render(request, 'arsip_tata/bundle_sync.html', {
    })

def digitalisasi(request, year):
    if not request.user.is_authenticated:
        return redirect('login')
        
    gendata = GenerateScriptDigitalizedView(request, year=year)
    gendata.gencontext()
    # print(gendata.context)
    return render(request=request, 
                  template_name=gendata.template_name, 
                  context=gendata.context)

class GenerateScriptDigitalizedView:
    def __init__(self, request, year) -> None:
        self.__year = year
        self.__request = request
        
    def gencontext(self):
        if self.__request.GET.get("search"):
            query = self.__request.GET.get("search")
            items = Item.objects.filter(Q(filesize__isnull=False, uploaded_date__year=self.year) & (Q(title__icontains=query)  | Q(bundle__description__icontains=query)))
        else:
            items = Item.objects.filter(filesize__isnull=False, uploaded_date__year=self.year)[:100]
        datalist = []
        for ke, item in enumerate(items):
            folder = str(item.bundle.yeardate)
            # breakpoint()
            path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, "-".join([str(item.bundle.yeardate), str(item.bundle.box.box_number), str(item.bundle.bundle_number), str(item.item_number)]) + ".pdf")
            pdffound = False
            # coverfilename = ""
            if exists(path):
                pdffound = True
            #     coverfilename = "{}_{}-{}-{}-{}.png".format(__package__.split('.')[1], item.bundle.yeardate, item.bundle.box.box_number, item.bundle.bundle_number, item.item_number)
            
            datalist.append({
                "box_number": item.bundle.box.box_number,
                "bundle_number": item.bundle.bundle_number,
                "doc_number": item.item_number,
                "bundle_code": item.bundle.code,
                "bundle_title": item.bundle.description,
                "bundle_year": item.bundle.year_bundle,
                "doc_description": item.title,
                "doc_count": item.total,
                "bundle_orinot": "", #item.bundle.orinot,
                "row_number": ke + 1,
                "pdffound": pdffound,
                "doc_id": item.id,
                "coverfilepath": item.cover.url,
                "filesize": item.filesize,
                "pagecount": item.page_count,
                "doc_uuid_id": item.id, #item.uuid_id,
                # "pdftmpfound": pdftmpfound,
                
        })

        self.__template_name = "arsip_tata/datalist_digitalized.html"
        if self.__request.method == 'POST':
            pass
        else:        
            self.__context = {
                                'appname': __package__.split('.')[1], 
                                'data': datalist,
                                'year': self.year,
                                'form': SearchDocByYear()
                            }
    @property
    def context(self):
        return self.__context
    @property
    def template_name(self):
        return self.__template_name
    @property
    def year(self):
        return self.__year
    
def get_data_digitalisasi(year):
    datalist = []    
    # bundles = Bundle.objects.filter(yeardate=year)
    # for bundle in bundles:
        # items = Item.objects.filter(Q(bundle_id=bundle.id) & Q(filesize__isnull=False))
    items = Item.objects.filter(uploaded_date__year=year)

    
    # if docs.count() != 0:
    data = []    
    for item in items:
        folder = str(item.yeardate)
        # filepath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
        filepath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, "-".join([str(item.bundle.yeardate), str(item.bundle.box.box_number), str(item.bundle.bundle_number), str(item.item_number)]) + ".pdf")
        # infotime=""
        infodate=""
        if exists(filepath):
            infotime = os.stat(filepath).st_mtime
            infodate = datetime.fromtimestamp(infotime).strftime('%d-%m-%Y')
        
        mdict = {
            'description': item.title,
            'doc_count':item.total,
            'doc_type': "Buku", #item.doc_type,
            'year': item.bundle.year_bundle,
            'date_scan': infodate,
            'doc_number': item.item_number,
            'title': item.bundle.description,
            'folder': folder,
            'box_number': item.bundle.box.box_number,
            'date_upload': item.uploaded_date.strftime('%d-%m-%Y'),
        }
        data.append(mdict)
        
        # if len(data) != 0:
        #     datalist.append({'title': bundle.description, 'folder':folder, 'box_number': bundle.box.box_number, 'data': data})
    return data

def create_dadigital_xls(datalist, sheet, year):
    centervh = Alignment(horizontal='center', vertical='center', wrap_text=True)
    leftvh = Alignment(horizontal='left', vertical='center', wrap_text=True)
    centerv = Alignment(vertical='center', wrap_text=True)
    wraptxt = Alignment(wrap_text=True)

    thin_border1 = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    thin_border2 = Border(bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
    thin_border3 = Border(top=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
    thin_border4 = Border(right=Side(style='thin'),left=Side(style='thin'))
    thin_border5 = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    thin_border6 = Border(top=Side(style='thin'))
    thin_border7 = Border(left=Side(style='thin'), top=Side(style='thin'))
    thin_border8 = Border(right=Side(style='thin'), top=Side(style='thin'))

    


    font_style1 = Font(name='Arial Narrow', size=10, bold=True)
    font_style2 = Font(name='Arial', size=8.5)
    font_style3 = Font(name='Arial', size=8.5, italic=True)
    color1 = PatternFill(start_color="c6d5f7", fill_type = "solid")
    # sheet.merge_cells('A1:G1')
    # sheet.merge_cells('A2:G2')
    aligncenter = Alignment(horizontal='center')
    headerfont = Font(name='Arial Narrow', size=14, bold=True)
    sheet['A1'] = f"DAFTAR ARSIP INAKTIF YANG DIALIH MEDIAKAN PADA TAHUN {year}"
    sheet['A1'].alignment = aligncenter
    sheet['A1'].font = headerfont
    sheet['A2'] = "UNIT PENGOLAH: BALAI WILAYAH SUNGAI MALUKU UTARA"
    sheet['A2'].alignment = aligncenter
    sheet['A2'].font = headerfont

    sheet.title = "DATA SUDAH ALIHMEDIA"
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 3
    sheet.column_dimensions['D'].width = 7
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 10

    sheet.merge_cells('A1:H1')
    sheet.merge_cells('A2:H2')


    for cell in sheet["7:7"]:
        cell.alignment = centervh
        cell.font = font_style1
    sheet.merge_cells('A7:A8')
    sheet.merge_cells('B7:B8')
    sheet.merge_cells('C7:D8')
    sheet.merge_cells('E7:E8')
    sheet.merge_cells('F7:F8')
    sheet.merge_cells('G7:G8')
    sheet.merge_cells('H7:H8')

    for i in range(1, 9):
        sheet.cell(row=9, column=i).value = i
        sheet.cell(row=9, column=i).alignment = centervh
        sheet.cell(row=9, column=i).border = thin_border1
        sheet.cell(row=9, column=i).font = Font(name='Arial Narrow', size=8, bold=True)
        sheet.cell(row=7, column=i).fill = color1
        sheet.cell(row=8, column=i).fill = color1
        sheet.cell(row=9, column=i).fill = color1
        
    sheet.row_dimensions[9].height = 9
    sheet.cell(row=7, column=1).value = "NOMOR"
    sheet.cell(row=7, column=1).border = thin_border1
    sheet.cell(row=8, column=1).border = thin_border1
    
    sheet.cell(row=7, column=2).value = "JENIS ARSIP"
    sheet.cell(row=7, column=2).border = thin_border1
    sheet.cell(row=8, column=2).border = thin_border1
    
    sheet.cell(row=7, column=3).value = "JUMLAH ARSIP"
    sheet.cell(row=7, column=3).border = thin_border1
    sheet.cell(row=8, column=3).border = thin_border1
    sheet.cell(row=7, column=4).border = thin_border1
    sheet.cell(row=8, column=4).border = thin_border1

    sheet.cell(row=7, column=5).value = "KURUN WAKTU"
    sheet.cell(row=7, column=5).border = thin_border1
    sheet.cell(row=8, column=5).border = thin_border1

    sheet.cell(row=7, column=6).value = "JENIS TINDAKAN ALIH MEDIA"
    sheet.cell(row=7, column=6).border = thin_border1
    sheet.cell(row=8, column=6).border = thin_border1

    sheet.cell(row=7, column=7).value = "KETERANGAN WAKTU PELAKSANAAN"
    sheet.cell(row=7, column=7).border = thin_border1
    sheet.cell(row=8, column=7).border = thin_border1

    sheet.cell(row=7, column=8).value = "LINK FILE"
    sheet.cell(row=7, column=8).border = thin_border1
    sheet.cell(row=8, column=8).border = thin_border1

    i = 10
    result = datalist
    no = 0
    for res in result:
        no += 1
        sheet['{}{}'.format('A', i)].value = no

        sheet['{}{}'.format('B', i)].value = " ".join([res["description"], res["title"]]).replace("\n"," ")
        sheet['{}{}'.format('B', i)].alignment = leftvh
        
        sheet['{}{}'.format('C', i)].value = res["doc_count"]
        sheet['{}{}'.format('C', i)].alignment = centervh
        
        sheet['{}{}'.format('D', i)].value = res["doc_type"]
        sheet['{}{}'.format('D', i)].alignment = centervh
        
        sheet['{}{}'.format('E', i)].value = res["year"]
        sheet['{}{}'.format('E', i)].alignment = centervh

        sheet['{}{}'.format('F', i)].value = 'SCAN'
        sheet['{}{}'.format('F', i)].alignment = centervh

        sheet['{}{}'.format('G', i)].value = res["date_upload"]
        sheet['{}{}'.format('G', i)].alignment = centervh
        sheet['{}{}'.format('A', i)].border = thin_border3
        sheet['{}{}'.format('B', i)].border = thin_border3
        sheet['{}{}'.format('C', i)].border = thin_border3
        sheet['{}{}'.format('D', i)].border = thin_border3
        sheet['{}{}'.format('E', i)].border = thin_border3
        sheet['{}{}'.format('F', i)].border = thin_border3
        sheet['{}{}'.format('G', i)].border = thin_border3

        filelocation = os.path.join(__package__.split('.')[1], res['folder'], str(res['box_number']), str(res['doc_number']) + ".pdf")
        # sheet['{}{}'.format('H', i)].fill = color1
        sheet['{}{}'.format('H', i)].value = '=HYPERLINK(CONCATENATE(CONFIG!A1, "{}")'.format(filelocation) + ', "LIHAT")'
        sheet['{}{}'.format('H', i)].border = thin_border1
        sheet['{}{}'.format('H', i)].font = Font(underline='single', bold=True, color="96251b")
        sheet['{}{}'.format('H', i)].alignment = centervh
        
        
        i += 1

    sheet['{}{}'.format('A', i)].border = thin_border6
    sheet['{}{}'.format('B', i)].border = thin_border6
    sheet['{}{}'.format('C', i)].border = thin_border6
    sheet['{}{}'.format('D', i)].border = thin_border6
    sheet['{}{}'.format('E', i)].border = thin_border6
    sheet['{}{}'.format('F', i)].border = thin_border6
    sheet['{}{}'.format('G', i)].border = thin_border6

def year_list_digital(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'arsip_tata/year_list_digital.html', {
        'years': Year.objects.all(),
    })
    
def show_year_digital(request):
    if not request.user.is_authenticated:
        return redirect('login')
    context = {
    }
    return render(request=request, template_name='arsip_tata/show_year_digital.html', context=context)

def report_digitalisasi(request, year):
    datalist = get_data_digitalisasi(year)
    # return HttpResponse(datalist)            
    filename = f"data_{__package__.split('.')[1]}_dadigitalisasi_{year}.xlsx"

    wb = Workbook()
    wb.create_sheet("CONFIG")
    sheet = wb["CONFIG"]
    sheet['A1'].value = os.path.join("D:", "media") + "\\"
    # wb.create_sheet("DATA SUDAH ALIHMEDIA")
    sheet = wb.active
    create_dadigital_xls(datalist=datalist, sheet=sheet, year=year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)    
    wb.save(response)
    return response


def transnotret_report(request):
    pdf = io.BytesIO()
    doc = SimpleDocTemplate(pdf, pagesize=A4)
    styles = getSampleStyleSheet()
    title = "Daftar Dokumen Keluar"

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height)
    template = PageTemplate(frames=[frame], id='mytemplate')

    doc.addPageTemplates([template])
    elements = []
    elements.append(Paragraph(title, styles['Title']))
    mydata = []
    mydata.append(("No", "Judul Dokumen", "Tanggal", "Nama / Satker"))
    c_width = [0.4*inch, 4*inch, 1*inch, 1.5*inch]
    
    style2 = getSampleStyleSheet()
    style2 = style2["BodyText"]
    style2.wordWrap = 'CJK'
    today = datetime.today().strftime('%d %B %Y')
    filename = f"dokumen_belum_kembali_{today.replace(' ', '_')}.pdf"
    details = TransDetail.objects.filter(date_return__isnull=True)

    no = 0
    for detail in details:
        if detail.date_return == None:  
            no += 1
            myset = (Paragraph(str(no), style2) , Paragraph(detail.item.title + "<br/>" + detail.item.bundle.description.replace("\n", "<br/>") + "<br/>" + "<strong>" + detail.item.codegen + "</strong>", style2), Paragraph(detail.trans.date_trans.strftime('%d %b %Y'), style2),  Paragraph(f"{detail.trans.customer.name} / {detail.trans.customer.organization}", style2))
            mydata.append(myset)
            

    mytable = Table(mydata, colWidths=c_width, hAlign='LEFT')
    mytable.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightblue),
                       ('INNERGRID', (0,0), (-1,-1), 0.25, colors.red),
                       ('FONTSIZE',(0,0),(-1,0),12),
                       ('FONTSIZE',(0,1),(-1,-1),8),
                       ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                       ('ALIGN', (0,0), (-1,0), 'CENTER'),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                       ('VALIGN',(0,0),(-1,-1),'TOP'),
                       ]))
    elements.append(Spacer(1, 10))
    elements.append(mytable)
    elements.append(Spacer(1, 32))
    rightal = ParagraphStyle(name="RightAl",alignment=TA_RIGHT)
    elements.append(Paragraph(f'Ternate, {today}', rightal))
    elements.append(Spacer(1, 32))
    elements.append(Paragraph('(Petugas Arsip)&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;', rightal))

    doc.build(elements)
    pdf.seek(0)
    response = HttpResponse(pdf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline;filename={filename}'
    return response

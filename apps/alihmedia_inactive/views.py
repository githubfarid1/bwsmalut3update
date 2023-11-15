from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from .models import Bundle, Doc, Department
from django.contrib import messages
import os
from django.db.models import Q
from os.path import exists
from django.conf import settings
import json
import sys
import time
from datetime import datetime, timedelta
import fitz
from django.contrib.auth.decorators import permission_required, user_passes_test
from django.http import JsonResponse
from .forms import UploadFileForm, DeletePdfFile, SearchQRCodeForm, ListDocByBox, DeleteDoc, SearchDoc, ExportForm
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import Group
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from django.db.models import Max

# from django_user_agents.utils import get_user_agent

def getmenu():
    return Department.objects.all()
def getdatabybox(box_number, link):
    d = Department.objects.get(link=link)
    depname = d.name
    folder = d.folder
    docs = Doc.objects.filter(bundle__department_id__exact=d.id, bundle__box_number__exact=box_number)
    # return HttpResponse(docs)
    boxdata = []
    for ke, doc in enumerate(docs):
        path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
        pdffound = False
        coverfilename = ""
        if exists(path):
            pdffound = True
            coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.bundle.box_number, doc.doc_number)
        filetmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc.id}.pdf")
        pdftmpfound = False
        if exists(filetmppath):
            pdftmpfound = True
        
        boxdata.append({
            "box_number": doc.bundle.box_number,
            "bundle_number": doc.bundle.bundle_number,
            "doc_number": doc.doc_number,
            "bundle_code": doc.bundle.code,
            "bundle_title": doc.bundle.title,
            "bundle_year": doc.bundle.year,
            "doc_description": doc.description,
            "doc_count": doc.doc_count,
            "bundle_orinot": doc.bundle.orinot,
            "row_number": ke + 1,
            "pdffound": pdffound,
            "doc_id": doc.id,
            "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
            "filesize": doc.filesize,
            "pagecount": doc.page_count,
            "doc_uuid_id": doc.uuid_id,
            "pdftmpfound": pdftmpfound,
        })
    return (boxdata, depname, folder)

def getdatabyfolder(folder):
    d = Department.objects.get(folder=folder)
    depname = d.name
    folder = d.folder
    docs = Doc.objects.filter(bundle__department_id__exact=d.id)
    # return HttpResponse(docs)
    boxdata = []
    for ke, doc in enumerate(docs):
        path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
        pdffound = False
        coverfilename = ""
        if exists(path):
            pdffound = True
            coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.bundle.box_number, doc.doc_number)
        filetmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc.id}.pdf")
        pdftmpfound = False
        if exists(filetmppath):
            pdftmpfound = True
        
        boxdata.append({
            "box_number": doc.bundle.box_number,
            "bundle_number": doc.bundle.bundle_number,
            "doc_number": doc.doc_number,
            "bundle_code": doc.bundle.code,
            "bundle_title": doc.bundle.title,
            "bundle_year": doc.bundle.year,
            "doc_description": doc.description,
            "doc_count": doc.doc_count,
            "bundle_orinot": doc.bundle.orinot,
            "row_number": ke + 1,
            "pdffound": pdffound,
            "doc_id": doc.id,
            "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
            "filesize": doc.filesize,
            "pagecount": doc.page_count,
            "doc_uuid_id": doc.uuid_id,
            "pdftmpfound": pdftmpfound,
        })
    return boxdata

def getdata(method, parquery, link):
    query = ""
    if method == "GET":
        query = parquery

    isfirst = True
    boxlist = []
    d = Department.objects.get(link=link)
    if query == None or query == '':
        docs = Doc.objects.filter(bundle__department_id__exact=d.id)
    else:
        docs = Doc.objects.filter(Q(bundle__department_id__exact=d.id) & (Q(description__icontains=query)  | Q(bundle__title__icontains=query) | Q(bundle__year__contains=query)))
    isfirst = True
    curbox_number = ""
    curbundle_number = ""
    # mlink = d.link.replace(__package__.split('.')[1] + "_", "")    
    for ke, doc in enumerate(docs):
        path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], d.folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
        pdffound = False
        filesize = 0
        pagecount = 0
        coverfilename = ""
        if exists(path):
            pdffound = True
            coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], d.folder, doc.bundle.box_number, doc.doc_number)
        filetmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc.id}.pdf")
        pdftmpfound = False
        if exists(filetmppath):
            pdftmpfound = True

        if isfirst:
            isfirst = False

            curbox_number = doc.bundle.box_number
            boxlist.append({
                "box_number": doc.bundle.box_number,
                "bundle_number": doc.bundle.bundle_number,
                "doc_number": doc.doc_number,
                "bundle_code": doc.bundle.code,
                "bundle_title": doc.bundle.title,
                "bundle_year": doc.bundle.year,
                "doc_description": doc.description,
                "doc_count": doc.doc_count,
                "bundle_orinot": doc.bundle.orinot,
                "row_number": ke + 1,
                "pdffound": pdffound,
                "doc_id": doc.id,
                "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
                "filesize": doc.filesize,
                "pagecount": doc.page_count,
                "doc_uuid_id": doc.uuid_id,
                "pdftmpfound": pdftmpfound,
            })
            continue
        if curbox_number == doc.bundle.box_number:
            box_number = ""
        else:
            box_number = doc.bundle.box_number
            curbox_number = doc.bundle.box_number
        
        if curbundle_number == doc.bundle.bundle_number:
            bundle_number = ""
            bundle_code = ""
            bundle_title = ""
            bundle_year = ""
            bundle_orinot = ""
        else:
            bundle_number = doc.bundle.bundle_number
            curbundle_number = doc.bundle.bundle_number
            bundle_code = doc.bundle.code
            bundle_title = doc.bundle.title
            bundle_year = doc.bundle.year
            bundle_orinot = doc.bundle.orinot
        
        doc_number = doc.doc_number
        doc_description = doc.description
        doc_count = doc.doc_count
        boxlist.append({
            "box_number": box_number,
            "bundle_number": bundle_number,
            "doc_number": doc_number,
            "bundle_code": bundle_code,
            "bundle_title": bundle_title,
            "bundle_year": bundle_year,
            "doc_description": doc_description,
            "doc_count": doc_count,
            "bundle_orinot": bundle_orinot,
            "row_number": ke + 1,
            "pdffound": pdffound,
            "doc_id": doc.id,
            "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
            "filesize": doc.filesize,
            "pagecount": doc.page_count,
            "doc_uuid_id": doc.uuid_id,
            "pdftmpfound": pdftmpfound,
        })
        
    isfirst = True
    rowbox = 0
    rowbundle = 0
    boxspan = 1
    bundlespan = 1      
    for ke, box in enumerate(boxlist):
        if isfirst:
            isfirst = False
            rowbox = ke
            rowbundle = ke
            boxspan = 1
            bundlespan = 1      
            continue
        if box['box_number'] == "":
            boxspan += 1
        else:
            boxlist[rowbox]['boxspan'] = boxspan
            boxspan = 1
            rowbox = ke

        if box['bundle_number'] == "":
            bundlespan += 1
        else:
            boxlist[rowbundle]['bundlespan'] = bundlespan
            bundlespan = 1
            rowbundle = ke

    # for last record
    if docs.count() != 0:
        boxlist[rowbox]['boxspan'] = boxspan
        boxlist[rowbundle]['bundlespan'] = bundlespan

    return boxlist
def summarydata(data):
    sumscan = 0
    listyear = []
    for d in data:
        if d['bundle_year'] is not None and d['bundle_year'].strip() != '':
            listyear.append(d['bundle_year'])
        if d['pdffound'] == True:
            sumscan += 1
    unyears = list(set(listyear))
    # tes = unyears.sort()
    unyears.sort()
    unyearstr = ", ".join(unyears)
    sumnotscan = len(data) - sumscan
    try:
        percent = sumscan / len(data) * 100
    except:
        percent = 0

    return (len(data), sumscan, sumnotscan, percent, unyearstr )
# @permission_required('apps_alihmedia_inactive.irigasi')

class GenerateScriptView_old:
    def __init__(self, funcname, request) -> None:
        self.__funcname = funcname
        self.__request = request

    def gencontext(self):
        test_group = Group.objects.get(name='arsip')
        if test_group in self.__request.user.groups.all():
            self.__template_name = "alihmedia_inactive/arsip_view.html"
            data = getdata(method=self.__request.method, parquery=self.__request.GET.get("search"), link=self.__funcname)
            summary = summarydata(data)
            self.__context = {
                "data": data,
                "link": self.__funcname,
                "totscan": summary[1],
                "totnotscan": summary[2],
                "totdata": summary[0],
                "percent": f"{summary[3]:.3f}",
                "years": summary[4],
                "menu": getmenu(),
                "appname":__package__.split('.')[1],
            }
        else:
            test_group = Group.objects.get(name='asesor')
            if test_group in self.__request.user.groups.all():
                self.__template_name = "alihmedia_inactive/asesor_view.html"
            else:
                self.__template_name = "alihmedia_inactive/guest_view.html"

            if self.__request.method == 'POST':
                form = ListDocByBox(self.__request.POST or None)
                if form.is_valid():
                    box_number = form.cleaned_data['box_number']
                    data = getdatabybox(box_number, self.__funcname)
                    boxdata = data[0]
                    depname = data[1]                
                    folder = data[2]
                    # return HttpResponse(next)    
                    self.__context = {'data':boxdata, 
                                        'depname':depname, 
                                        'box_number': box_number, 
                                        "folder": folder, 
                                        'form': ListDocByBox(),
                                        "menu": getmenu(), 
                                        "appname":__package__.split('.')[1], 
                                        "link": self.__funcname}
            else:        
                self.__context = {'form':ListDocByBox(), 
                                    'menu': getmenu(), 
                                    'appname': __package__.split('.')[1], 
                                    'link': self.__funcname}
    @property
    def context(self):
        return self.__context
    @property
    def template_name(self):
        return self.__template_name

class GenerateScriptView:
    def __init__(self, funcname, request) -> None:
        self.__funcname = funcname
        self.__request = request

    def gencontext(self):
        self.__template_name = "alihmedia_inactive/datalist.html"
        if self.__request.method == 'POST':
            form = ListDocByBox(self.__request.POST or None)
            if form.is_valid():
                box_number = form.cleaned_data['box_number']
                data = getdatabybox(box_number, self.__funcname)
                boxdata = data[0]
                depname = data[1]                
                folder = data[2]
                self.__context = {'data':boxdata, 
                                    'depname':depname, 
                                    'box_number': box_number, 
                                    "folder": folder, 
                                    'form': ListDocByBox(),
                                    "menu": getmenu(), 
                                    "appname":__package__.split('.')[1], 
                                    "link": self.__funcname}
        else:        
            self.__context = {'form':ListDocByBox(), 
                                'menu': getmenu(), 
                                'appname': __package__.split('.')[1], 
                                'link': self.__funcname}
    @property
    def context(self):
        return self.__context
    @property
    def template_name(self):
        return self.__template_name

@csrf_exempt
def irigasi(request):
    if not request.user.is_authenticated:
        return redirect('login')
    data = GenerateScriptView(__package__.split('.')[1] + "_" + sys._getframe().f_code.co_name, request)
    data.gencontext()
    return render(request=request, template_name=data.template_name, context=data.context)

@csrf_exempt
def air_baku(request):
    if not request.user.is_authenticated:
        return redirect('login')
    data = GenerateScriptView(__package__.split('.')[1] + "_" + sys._getframe().f_code.co_name, request)
    data.gencontext()
    return render(request=request, template_name=data.template_name, context=data.context)

@csrf_exempt
def sungai(request):
    if not request.user.is_authenticated:
        return redirect('login')
    data = GenerateScriptView(__package__.split('.')[1] + "_" + sys._getframe().f_code.co_name, request)
    data.gencontext()
    return render(request=request, template_name=data.template_name, context=data.context)

@csrf_exempt
def pantai(request):
    if not request.user.is_authenticated:
        return redirect('login')
    data = GenerateScriptView(__package__.split('.')[1] + "_" + sys._getframe().f_code.co_name, request)
    data.gencontext()
    return render(request=request, template_name=data.template_name, context=data.context)

@csrf_exempt
def keuangan(request):
    if not request.user.is_authenticated:
        return redirect('login')
    data = GenerateScriptView(__package__.split('.')[1] + "_" + sys._getframe().f_code.co_name, request)
    data.gencontext()
    return render(request=request, template_name=data.template_name, context=data.context)

@user_passes_test(lambda user: Group.objects.get(name='arsip') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all())
def pdfdownload(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')

    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.bundle.department.folder
    box_number = doc.bundle.box_number
    doc_number = doc.doc_number
    # link = link.replace(__package__.split('.')[1] + "_", "")
    
    path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(box_number), str(doc_number) + ".pdf")
    if exists(path):
        filename = f"{__package__.split('.')[1]}_{folder}_{box_number}_{doc_number}.pdf"
        with open(path, 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline;filename={filename}'
            return response
    raise Http404

def statistics(request):
    if not request.user.is_authenticated:
        return redirect('login')

    deps = Department.objects.all()
    depnamelist = []
    depvaluelist = []
    colorlist = []
    foundall = 0
    notfoundall = 0
    for d in deps:
        folder = d.folder
        foundlist = [os.path.join(root, file) for root, dirs, files in os.walk(os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder)) for file in files if file.endswith(".pdf")]
        found = len(foundlist)
        foundall += found
        docs = Doc.objects.filter(bundle__department_id__exact=d.id)
        notfound = len(docs) - found
        notfoundall += notfound
        depnamelist.append(" | ".join([d.name, "Sudah"]))
        depnamelist.append(" | ".join([d.name, "Belum"]))
        depvaluelist.append(found)
        depvaluelist.append(notfound)
        colorlist.append("rgba(112, 185, 239, 1)")
        colorlist.append("rgba(244, 204, 204, 1)")
        total = foundall + notfoundall
        procfound = foundall / total * 100
        procnotfound = notfoundall / total * 100
    
    fileinfolist = []
    allfilelist = [os.path.join(root, file) for root, dirs, files in os.walk(os.path.join(settings.PDF_LOCATION, __package__.split('.')[1])) for file in files if file.endswith(".pdf")]
    for filepath in allfilelist:
        # infotime = os.path.getmtime(filepath)
        infotime = os.stat(filepath).st_mtime
        infodate = datetime.fromtimestamp(infotime).strftime('%d-%m-%Y')
        mdict = {
            "file": filepath,
            "date": infodate,
            "pages": fitz.open(filepath).page_count
        }
        fileinfolist.append(mdict)

    num_of_dates = 30
    start = datetime.today()
    date_list = [start.date() - timedelta(days=x) for x in range(num_of_dates)]
    date_list.sort()
    docscan = []
    doccolor = []
    docdate = []
    # print(date_list)
    for d in date_list:
        pages = 0
        for fl in fileinfolist:
            if fl['date'] == d.strftime('%d-%m-%Y'):
                pages += fl['pages']
        docdate.append(d.strftime('%d-%m-%Y'))
        docscan.append(pages)
        doccolor.append("rgba(112, 185, 239, 1)")
    context = {
        "menu": getmenu(),
        "depnamelist": depnamelist,
        "depvaluelist": depvaluelist,
        "colorlist": colorlist,
        "foundall": str(foundall),
        "notfoundall": str(notfoundall),
        "procfound":f"{procfound:.3f}",
        "procnotfound":f"{procnotfound:.3f}",
        "docdate": docdate,
        "docscan": docscan,
        "doccolor": doccolor,
    }
    return render(request=request, template_name='alihmedia_inactive/statistics.html', context=context)

@csrf_exempt
def boxsearch(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == 'POST':
        qrcode = request.POST.get("qrcode")
        strlist = qrcode.strip().split('/')
        try:
            folder = strlist[3]
            box_number = strlist[4]
        except:
            messages.info(request, "Data tidak ditemukan")
            context = {}
            context['form'] = SearchQRCodeForm()
            return render(request, 'alihmedia_inactive/boxsearch2.html', context=context)
            # return HttpResponse("QRcode Error")
        
        # return HttpResponse(folder + box)
        d = Department.objects.get(folder=folder)
        depname = d.name
        docs = Doc.objects.filter(bundle__department_id__exact=d.id, bundle__box_number__exact=box_number)
        boxdata = []
        for ke, doc in enumerate(docs):
            path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
            pdffound = False
            coverfilename = ""
            if exists(path):
                pdffound = True
                coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.bundle.box_number, doc.doc_number)
            boxdata.append({
                "box_number": doc.bundle.box_number,
                "bundle_number": doc.bundle.bundle_number,
                "doc_number": doc.doc_number,
                "bundle_code": doc.bundle.code,
                "bundle_title": doc.bundle.title,
                "bundle_year": doc.bundle.year,
                "doc_description": doc.description,
                "doc_count": doc.doc_count,
                "bundle_orinot": doc.bundle.orinot,
                "row_number": ke + 1,
                "pdffound": pdffound,
                "doc_id": doc.id,
                "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
                "filesize": doc.filesize,
                "pagecount": doc.page_count,
                "doc_uuid_id": doc.uuid_id,
            })

        # return HttpResponse(docs[2].bundle.title)
        # context['form'] = SearchQRCodeForm()
        context = {'data':boxdata, 'depname':depname, 'box_number': box_number, "folder": folder, 'form': SearchQRCodeForm()}
        return render(request=request, template_name='alihmedia_inactive/boxsearch2.html', context=context)
        # pass
    context = {}
    context['form'] = SearchQRCodeForm()
    # context['url'] = url
    return render(request, 'alihmedia_inactive/boxsearch2.html', context=context)

@csrf_exempt
def pdfupload(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')
    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.bundle.department.folder
    doc_id = doc.id
    pdfpath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
    tmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc_id}.pdf")
    if exists(pdfpath):
        messages.info(request, "File sudah ada")
        return redirect(f"/{__package__.split('.')[1]}/{folder}#{str(doc.bundle.box_number)}")


    if request.method == 'POST' and request.FILES['filepath'] and not exists(pdfpath):
        upload = request.FILES['filepath']
        fss = FileSystemStorage()
        if exists(tmppath):
            os.remove(tmppath)
        fss.save(tmppath, upload)
        messages.info(request, "File berhasil diupload, akan segera diproses")
        # time.sleep(2)
        return redirect(f"/{__package__.split('.')[1]}/{folder}#{str(doc.bundle.box_number)}")

    context = {}
    context['form'] = UploadFileForm(initial={'uuid_id': uuid_id})
    context['isexist'] = exists(tmppath)
    context['data'] = doc

    # context['url'] = url
    return render(request,'alihmedia_inactive/pdfupload.html', context=context)

@user_passes_test(lambda user: Group.objects.get(name='arsip') in user.groups.all())
@csrf_exempt
def pdfremove(request, uuid_id):
    if not request.user.is_authenticated:
        return redirect('login')
    doc = Doc.objects.get(uuid_id=uuid_id)
    folder = doc.bundle.department.folder
    doc_id = doc.id
    pdfpath = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
    if not exists(pdfpath):
        messages.info(request, "File tidak ada")
        return redirect(f"/{__package__.split('.')[1]}/{folder}#{str(doc.bundle.box_number)}")
    coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.bundle.box_number, doc.doc_number)
    if request.method == 'POST':
        if exists(pdfpath):
            # ts = str(time.time())
            # pdfrename = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf." + str(ts))
            # os.rename(pdfpath, pdfrename)
            os.remove(pdfpath)
            coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], folder, doc.bundle.box_number, doc.doc_number)
            if exists(os.path.join(settings.COVER_LOCATION, coverfilename)):
                os.remove(os.path.join(settings.COVER_LOCATION, coverfilename))
            tmppath = os.path.join(settings.MEDIA_ROOT, "tmpfiles", f"{__package__.split('.')[1]}-{doc_id}.pdf")
            if exists(tmppath):
                os.remove(tmppath)
            doc.filesize = None
            doc.page_count = None
            doc.save()     
            messages.info(request, "Berhasil dihapus")
            return redirect(f"/{__package__.split('.')[1]}/{folder}#{doc.bundle.box_number}")
        else:
            messages.info(request, "File tidak ada")
    context = {}
    context['uuid_id'] = uuid_id
    context['isexist'] = exists(pdfpath)
    context['data'] = doc
    context["coverfilepath"] =  os.path.join(settings.COVER_URL, coverfilename)

    # context['url'] = url
    return render(request,'alihmedia_inactive/pdfremove.html', context=context)

def searchdoc(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.GET.get("folder"):
        query = request.GET.get("search")
        folder = request.GET.get("folder")
        d = Department.objects.get(folder=folder)
        docs = Doc.objects.filter(Q(bundle__department_id__exact=d.id) & (Q(description__icontains=query)  | Q(bundle__title__icontains=query) | Q(bundle__year__contains=query)))
        if not docs:
            messages.info(request, "Data tidak ditemukan")
        boxdata = []
        for ke, doc in enumerate(docs):
            path = os.path.join(settings.PDF_LOCATION, __package__.split('.')[1], doc.bundle.department.folder, str(doc.bundle.box_number), str(doc.doc_number) + ".pdf")
            pdffound = False
            filesize = 0
            pagecount = 0
            coverfilename = ""
            if exists(path):
                pdffound = True
                coverfilename = "{}_{}_{}_{}.png".format(__package__.split('.')[1], doc.bundle.department.folder, doc.bundle.box_number, doc.doc_number)
            boxdata.append({
                "department_folder": doc.bundle.department.folder,
                "box_number": doc.bundle.box_number,
                "bundle_number": doc.bundle.bundle_number,
                "doc_number": doc.doc_number,
                "bundle_code": doc.bundle.code,
                "bundle_title": doc.bundle.title,
                "bundle_year": doc.bundle.year,
                "doc_description": doc.description,
                "doc_count": doc.doc_count,
                "bundle_orinot": doc.bundle.orinot,
                "row_number": ke + 1,
                "pdffound": pdffound,
                "doc_id": doc.id,
                "coverfilepath": os.path.join(settings.COVER_URL, coverfilename),
                "filesize": doc.filesize,
                "pagecount": doc.page_count,
                "doc_uuid_id": doc.uuid_id,
            })
        context = {'data':boxdata, 'form': SearchDoc(), 'folder':folder, 'query':query}
        return render(request,'alihmedia_inactive/searchdoc.html', context=context)

    context = {}
    context['form'] = SearchDoc()
    return render(request,'alihmedia_inactive/searchdoc.html', context=context)

# def add(request):
#     if not request.user.is_authenticated:
#         return redirect('login')
#     addForm=DocAddForm()
#     if request.method=='POST':
#         docAdd=DocAddForm(request.POST)
#         folder = request.POST.get('folder')
#         variety = Variety.objects.get(folder=folder)
#         docmax = Doc.objects.filter(variety=variety).aggregate(Max('doc_number'))
#         if docAdd.is_valid():
#             instance = docAdd.save(commit=False)
#             instance.variety_id = variety.id
#             instance.doc_number = docmax['doc_number__max'] + 1
#             instance.save()
#             messages.success(request, 'Data has been added.')
#             next = request.POST.get('next', '/')
#             return redirect(next)
#     return render(request,'alihmedia_vital/add.html',{
#         'form':addForm})

def create_xls(datalist, app_name, folder):
    wb = Workbook()
    wb.create_sheet("CONFIG")
    sheet = wb["CONFIG"]
    sheet['A1'].value = os.path.join("D:", "media") + "\\"

    sheet = wb.active
    sheet.title = "DATA INAKTIF"
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['B'].width = 7
    sheet.column_dimensions['C'].width = 7
    sheet.column_dimensions['D'].width = 7
    sheet.column_dimensions['E'].width = 57
    sheet.column_dimensions['F'].width = 45
    sheet.column_dimensions['G'].width = 7
    sheet.column_dimensions['H'].width = 5 
    sheet.column_dimensions['I'].width = 7 
    sheet.column_dimensions['J'].width = 8
    sheet.merge_cells('A1:J1')
    sheet['A1'] = "DAFTAR ARSIP INAKTIF"
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].font = Font(name='Arial Narrow', size=14, bold=True)
    centervh = Alignment(horizontal='center', vertical='center', wrap_text=True)
    centerv = Alignment(vertical='center', wrap_text=True)
    wraptxt = Alignment(wrap_text=True)

    thin_border1 = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    thin_border2 = Border(bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
    thin_border3 = Border(top=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
    thin_border4 = Border(right=Side(style='thin'),left=Side(style='thin'))
    thin_border5 = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    font_style1 = Font(name='Arial Narrow', size=11, bold=True)
    font_style2 = Font(name='Arial', size=8.5)
    font_style3 = Font(name='Arial', size=8.5, italic=True)
    color1 = PatternFill(start_color="c6d5f7", fill_type = "solid")

    sheet.row_dimensions[7].height = 28
    for cell in sheet["7:7"]:
        cell.alignment = centervh
        cell.font = font_style1
    
    headers = ("NO BOX", "NO BRKS", "NO URUT", "KODE", "INDEKS", "URAIAN MASALAH", "THN", "JML", "KET", "ACTION")
    for i in headers:
        sheet.cell(row=7, column=headers.index(i)+1).value = i
        sheet.cell(row=7, column=headers.index(i)+1).border = thin_border1
    
    i = 8
    result = datalist
    curbox = result[0]["box_number"]
    curbundle = result[0]["bundle_number"]
    isfirst = True
    for res in result:
        sheet['{}{}'.format('A', i)].border = thin_border4
        sheet['{}{}'.format('B', i)].border = thin_border4
        sheet['{}{}'.format('C', i)].border = thin_border4
        sheet['{}{}'.format('D', i)].border = thin_border4
        sheet['{}{}'.format('E', i)].border = thin_border4
        sheet['{}{}'.format('F', i)].border = thin_border4
        sheet['{}{}'.format('G', i)].border = thin_border4
        sheet['{}{}'.format('H', i)].border = thin_border4
        sheet['{}{}'.format('I', i)].border = thin_border4
        sheet['{}{}'.format('J', i)].border = thin_border4
        sheet['{}{}'.format('C', i)].border = thin_border5
        sheet['{}{}'.format('F', i)].border = thin_border5
        sheet['{}{}'.format('H', i)].border = thin_border5
        if isfirst:
            isfirst = False
            sheet['{}{}'.format('A', i)].value = res["box_number"]
            sheet['{}{}'.format('B', i)].value = res["bundle_number"]
            sheet['{}{}'.format('C', i)].value = res["doc_number"]
            sheet['{}{}'.format('D', i)].value = res["bundle_code"]
            sheet['{}{}'.format('E', i)].value = res["bundle_title"]
            sheet['{}{}'.format('F', i)].value = res["doc_description"]
            sheet['{}{}'.format('G', i)].value = res["bundle_year"]
            sheet['{}{}'.format('H', i)].value = res["doc_count"]
            sheet['{}{}'.format('I', i)].value = res["bundle_orinot"]
        else:
            if curbox != res["box_number"]:
                curbox = res["box_number"]
                sheet['{}{}'.format('A', i)].value = res["box_number"]
                sheet['{}{}'.format('A', i)].border = thin_border3
            else:
                sheet['{}{}'.format('A', i)].value = "" 
            if curbundle != res["bundle_number"]:
                curbundle = res["bundle_number"]
                sheet['{}{}'.format('B', i)].value = res["bundle_number"]
                sheet['{}{}'.format('D', i)].value = res["bundle_code"]
                sheet['{}{}'.format('E', i)].value = res["bundle_title"]
                sheet['{}{}'.format('G', i)].value = res["bundle_year"]
                sheet['{}{}'.format('I', i)].value = res["bundle_orinot"]

                sheet['{}{}'.format('B', i)].border = thin_border3
                sheet['{}{}'.format('D', i)].border = thin_border3
                sheet['{}{}'.format('E', i)].border = thin_border3
                sheet['{}{}'.format('G', i)].border = thin_border3
                sheet['{}{}'.format('I', i)].border = thin_border3

            else:
                sheet['{}{}'.format('B', i)].value = ""
                sheet['{}{}'.format('D', i)].value = ""
                sheet['{}{}'.format('E', i)].value = ""
                sheet['{}{}'.format('G', i)].value = ""
                sheet['{}{}'.format('I', i)].value = ""
            sheet['{}{}'.format('C', i)].value = res["doc_number"]
            sheet['{}{}'.format('F', i)].value = res["doc_description"]
            sheet['{}{}'.format('H', i)].value = res["doc_count"]
        # sheet.cell(row=row, column=1).alignment = centerv
        sheet['{}{}'.format('A', i)].alignment = centervh
        sheet['{}{}'.format('B', i)].alignment = centervh
        sheet['{}{}'.format('C', i)].alignment = centervh
        sheet['{}{}'.format('D', i)].alignment = centervh
        sheet['{}{}'.format('E', i)].alignment = centerv
        sheet['{}{}'.format('F', i)].alignment = centerv
        sheet['{}{}'.format('G', i)].alignment = centervh
        sheet['{}{}'.format('H', i)].alignment = centervh
        sheet['{}{}'.format('I', i)].alignment = centervh
        sheet['{}{}'.format('J', i)].alignment = centervh
        
        if res['filesize'] is not None:
            filelocation = os.path.join(app_name, folder, str(res['box_number']), str(res['doc_number']) + ".pdf")
            sheet['{}{}'.format('F', i)].fill = color1
            sheet['{}{}'.format('J', i)].fill = color1
            sheet['{}{}'.format('J', i)].value = '=HYPERLINK(CONCATENATE(CONFIG!A1, "{}")'.format(filelocation) + ', "LIHAT")'
            sheet['{}{}'.format('J', i)].border = thin_border1
            sheet['{}{}'.format('J', i)].font = Font(underline='single', bold=True, color="96251b")
        i += 1
    return wb

@user_passes_test(lambda user: Group.objects.get(name='arsip') in user.groups.all() or Group.objects.get(name='asesor') in user.groups.all())
def export(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.GET.get("folder"):
        folder = request.GET.get("folder")
        datalist = getdatabyfolder(folder)
        # return HttpResponse(json.dumps(datalist, default=str), content_type="application/json")
        filename = f"data_{__package__.split('.')[1]}_{folder}.xlsx"
        wb = create_xls(datalist, __package__.split('.')[1], folder)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)    
        wb.save(response)
        return response
    context = {}
    context['form'] = ExportForm()
    return render(request=request, template_name='alihmedia_inactive/export.html', context=context)

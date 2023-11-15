from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
wb = Workbook()
sheet = wb.active
 
sheet.title = "DATA VITAL"
sheet.column_dimensions['A'].width = 4.1
sheet.column_dimensions['B'].width = 37
sheet.column_dimensions['C'].width = 11.3
sheet.column_dimensions['D'].width = 8.8
sheet.column_dimensions['E'].width = 9.5
sheet.column_dimensions['F'].width = 12.8
sheet.column_dimensions['G'].width = 16
sheet.column_dimensions['H'].width = 23.5 
sheet.column_dimensions['I'].width = 15.5 
sheet.column_dimensions['J'].width = 14.6
sheet.merge_cells('A1:J1')
sheet['A1'] = "DAFTAR ARSIP VITAL"
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['A1'].font = Font(name='Arial', size=12, bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
sheet.row_dimensions[7].height = 28
for cell in sheet["7:7"]:
    cell.alignment = center
    cell.font = Font(name='Arial', size=8.5, bold=True)
headers = ("NO", "JENIS ARSIP", "UNIT KERJA", "KURUN WAKTU", "MEDIA", "JUMLAH", "JANGKA SIMPAN", "LOKASI SIMPAN", "METODE PERLINDUNGAN", "KETERANGAN")
for i in headers:
    sheet.cell(row=7, column=headers.index(i)+1).value = i
    sheet.cell(row=7, column=headers.index(i)+1).border = thin_border


wb.save("test.xlsx")
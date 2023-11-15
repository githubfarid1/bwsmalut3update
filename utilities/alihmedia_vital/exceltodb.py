from openpyxl import Workbook, load_workbook
import json
import argparse
import sys
import os
from settings import *
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session
from dbclass import Doc, Variety, Base
import uuid
import contextlib

engine = create_engine('mysql+pymysql://{}:{}@localhost:{}/{}'.format(USER, PASSWORD, PORT, DBNAME) , echo=False)


def getbegin(ws, rownum, col):
    while True:
        # print(ws[f"{col}{rownum}"].value)
        if ws[f"{col}{rownum}"].value == None:
            rownum += 1
        else:
            return rownum


def checkexcel(ws, sheetname):
    for i in range(7, ws.max_row + 1):
        if ws[f"C{i}"].value is None and ws[f"G{i}"].value is not None:
            print("check "+ sheetname + " row: " + str(i) + " " + str(ws[f"G{i}"].value))

def parse(ws):
    # [{"variety":"BPKB Mobil dan Motor", "data": [{"name": "BPKB Toyota / Kijang KF40 (B 8525 DR)", "unit_kerja":"xxxx"}, {"name": "BPKB Toyota / Kijang KF40 (B 8525 DR)", "unit_kerja":"xxxx"}]},
    # {"variety":"BPKB Mobil dan Motor", "data": [{"name": "BPKB Toyota / Kijang KF40 (B 8525 DR)", "unit_kerja":"xxxx"}, {"name": "BPKB Toyota / Kijang KF40 (B 8525 DR)", "unit_kerja":"xxxx"}]}
    # ]
    datalist = []
    ke = -1
    for i in range(8, ws.max_row + 1):
        mdict = {}
        if ws[f"C{i}"].value is None:
            ke += 1
            mdict["jenis"] = ws[f"B{i}"].value
            mdict["data"] = []
            datalist.append(mdict)
            continue
        else:
            datalist[ke]["data"].append({"nama": ws[f"B{i}"].value, "unit_kerja": ws[f"C{i}"].value, "kurun_waktu": ws[f"D{i}"].value, "media": ws[f"E{i}"].value, "jumlah": ws[f"F{i}"].value, "jangka_simpan": ws[f"G{i}"].value, "lokasi_simpan": ws[f"H{i}"].value, "metode_proteksi": ws[f"I{i}"].value, "keterangan": ws[f"J{i}"].value, "nomor": ws[f"A{i}"].value})
    return datalist

def listtodb(datalist, session):
    for data in datalist:
        variety = Variety(name=data['jenis'], folder=str(data['jenis']).lower().replace(" ", "_") )
        session.add(variety)
        session.flush()
        session.commit()
        varietyId = variety.id
        for d in data['data']:
            doc = Doc(name=d['nama'], variety_id=varietyId, work_unit=d['unit_kerja'], period=d['kurun_waktu'], media=d['media'], countstr=d['jumlah'], save_life=d['jangka_simpan'], uuid_id=uuid.uuid4().hex, save_location=d['lokasi_simpan'], protect_method=d['metode_proteksi'], description=d['keterangan'], doc_number=d['nomor'])
            session.add(doc)
        session.flush()
        session.commit()



def main():
    #TRUNCATE TABLES
    with contextlib.closing(engine.connect()) as con:
        trans = con.begin()
        for table in reversed(Base.metadata.sorted_tables):
            con.execute(table.delete())
            con.execute(text('ALTER TABLE %s AUTO_INCREMENT = 1' % table.name))
        trans.commit()    

    #START
    session = Session(engine)
    wb = load_workbook(filename=EXCEL_FILE, data_only=True)
    ws = wb[EXCEL_SHEET]
    datalist = parse(ws)

    listtodb(datalist=datalist, session=session)
    print("Success")

if __name__ == '__main__':
    main()

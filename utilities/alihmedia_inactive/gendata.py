from openpyxl import Workbook, load_workbook
import json

fname = "Daftar Arsip.xlsx"
wb = load_workbook(filename=fname)
ws = wb["IRIGASI"]
def getbegin(rownum, col):
    while True:
        if ws[f"{col}{rownum}"].value == None:
            rownum += 1
        else:
            return rownum

boxlist = []
first = True
for i in range(7, ws.max_row + 1):
    if first:
        begin = i
        first = False
        boxno = ws[f'A{i}'].value
        continue

    if ws[f"A{i}"].value != None:
        # end = getend(i)
        end = i-1
        boxlist.append({"box": boxno, "begin": begin, "end": end})

        begin = getbegin(i, "C")
        boxno = ws[f'A{i}'].value

boxlist.append({ "box": boxno, "begin": begin, "end": i-1})

for ke, box in enumerate(boxlist):
    noberkas = ws[f"B{box['begin']}"].value
    first = True
    berkaslist = []
    for i in range(box['begin'], box['end']+1):
        if first:
            begin = box['begin']
            first = False
            berkasno = ws[f'B{i}'].value
            continue

        if ws[f"B{i}"].value != None and  str(ws[f"B{i}"].value).strip() != "":
            end = i-1
            index = ""
            perus = ""
            nourutlist = []
            for i2 in range(begin, end):
                if ws[f"E{i2}"].value != None and str(ws[f"E{i2}"].value).strip() != "":
                    try:
                        index = index + " " + ws[f"E{i2}"].value
                    except:
                        pass

                if ws[f"F{i2}"].value != None and str(ws[f"F{i2}"].value).strip() != "":
                    try:
                        perus = perus + " " + ws[f"F{i2}"].value
                    except:
                        pass
                if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
                    try:
                        nourutlist.append({"nourut": ws[f"C{i2}"].value, "uraian": ws[f"G{i2}"].value, "jumlah": ws[f"I{i2}"].value })
                    except:
                        pass

            dtemp = {"berkas": berkasno, "kode": ws[f"D{begin}"].value, "tahun": str(ws[f"H{begin}"].value), "ket": str(ws[f"J{begin}"].value), "index": perus.strip() + "\n" + index.strip(), "begin": begin, "end": end, "data": nourutlist}
            berkaslist.append(dtemp)
            berkasno = ws[f"B{i}"].value
            begin = getbegin(i, "C")

    #DATA BERKAS TERBAWAH
    index = ""
    perus = ""
    nourutlist = []
    for i2 in range(begin, box['end']+1):
        if ws[f"E{i2}"].value != None and str(ws[f"E{i2}"].value).strip() != "":
            try:
                index = index + " " + ws[f"E{i2}"].value
            except:
                pass

        if ws[f"F{i2}"].value != None and str(ws[f"F{i2}"].value).strip() != "":
            try:
                perus = perus + " " + ws[f"F{i2}"].value
            except:
                pass

        if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
            try:
                nourutlist.append({"nourut": ws[f"C{i2}"].value, "uraian": ws[f"G{i2}"].value, "jumlah": ws[f"I{i2}"].value })
            except:
                pass

    dtemp = {"berkas": berkasno, "kode": ws[f"D{begin}"].value, "tahun": str(ws[f"H{begin}"].value), "ket": str(ws[f"J{begin}"].value), "index": perus.strip() + "\n" + index.strip(), "begin":begin, "end": box['end'], "data": nourutlist}
    berkaslist.append(dtemp)
    boxlist[ke]['data'] = berkaslist

with open("data.json", 'w') as file:
    json.dump(boxlist, file)



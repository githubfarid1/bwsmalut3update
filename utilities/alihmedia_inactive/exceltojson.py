from openpyxl import Workbook, load_workbook
import json
import argparse
import sys
import os
from settings import *
def getbegin(ws, rownum, col):
    while True:
        # print(ws[f"{col}{rownum}"].value)
        if ws[f"{col}{rownum}"].value == None:
            rownum += 1
        else:
            return rownum

def parse(ws):
    boxlist = []
    first = True
    for i in range(7, ws.max_row + 1):
        if first:
            begin = i
            first = False
            boxno = ws[f'A{i}'].value
            continue

        if ws[f"A{i}"].value != None:
            # end = getend(i)
            end = i-1
            boxlist.append({"box": boxno, "begin": begin, "end": end})

            begin = getbegin(ws, i, "C")
            boxno = ws[f'A{i}'].value

    boxlist.append({ "box": boxno, "begin": begin, "end": i-1})

    for ke, box in enumerate(boxlist):
        noberkas = ws[f"B{box['begin']}"].value
        first = True
        berkaslist = []
        for i in range(box['begin'], box['end']+1):
            if first:
                begin = box['begin']
                first = False
                berkasno = ws[f'B{i}'].value
                continue

            if ws[f"B{i}"].value != None and  str(ws[f"B{i}"].value).strip() != "":
                end = i-1
                index = ""
                perus = ""
                nourutlist = []
                for i2 in range(begin, end):
                    if ws[f"E{i2}"].value != None and str(ws[f"E{i2}"].value).strip() != "":
                        try:
                            index = index + " " + ws[f"E{i2}"].value
                        except:
                            pass

                    if ws[f"F{i2}"].value != None and str(ws[f"F{i2}"].value).strip() != "":
                        try:
                            perus = perus + " " + ws[f"F{i2}"].value
                        except:
                            pass
                    if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
                        try:
                            nourutlist.append({"nourut": ws[f"C{i2}"].value, "uraian": ws[f"G{i2}"].value, "jumlah": ws[f"I{i2}"].value })
                        except:
                            pass

                dtemp = {"berkas": berkasno, "kode": ws[f"D{begin}"].value, "tahun": str(ws[f"H{begin}"].value), "ket": str(ws[f"J{begin}"].value), "index": perus.strip() + "\n" + index.strip(), "begin": begin, "end": end, "data": nourutlist}
                berkaslist.append(dtemp)
                berkasno = ws[f"B{i}"].value
                begin = getbegin(ws, i, "C")

        #DATA BERKAS TERBAWAH
        index = ""
        perus = ""
        nourutlist = []
        for i2 in range(begin, box['end']+1):
            if ws[f"E{i2}"].value != None and str(ws[f"E{i2}"].value).strip() != "":
                try:
                    index = index + " " + ws[f"E{i2}"].value
                except:
                    pass

            if ws[f"F{i2}"].value != None and str(ws[f"F{i2}"].value).strip() != "":
                try:
                    perus = perus + " " + ws[f"F{i2}"].value
                except:
                    pass

            if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
                try:
                    nourutlist.append({"nourut": ws[f"C{i2}"].value, "uraian": ws[f"G{i2}"].value, "jumlah": ws[f"I{i2}"].value })
                except:
                    pass

        dtemp = {"berkas": berkasno, "kode": ws[f"D{begin}"].value, "tahun": str(ws[f"H{begin}"].value), "ket": str(ws[f"J{begin}"].value), "index": perus.strip() + "\n" + index.strip(), "begin":begin, "end": box['end'], "data": nourutlist}
        berkaslist.append(dtemp)
        boxlist[ke]['data'] = berkaslist
    return boxlist

def main():
    parser = argparse.ArgumentParser(description="Get data from excel and save them to json")
    # parser.add_argument('-input', '--xlsinput', type=str,help="XLSX File Input")
    # parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    parser.add_argument('-output', '--jsonoutput', type=str,help="File output in json")

    args = parser.parse_args()
    # if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
    #     input('input the right XLSX or XLSM file')
    #     sys.exit()

    # isExist = os.path.exists(args.xlsinput)
    # if not isExist:
    #     input(args.xlsinput + " does not exist")
    #     sys.exit()

    # fname = "Daftar Arsip.xlsx"
    # wb = load_workbook(filename=fname)
    # ws = wb["IRIGASI"]
    mainboxlist = []
    wb = load_workbook(filename=EXCEL_FILE)
    for sheetname in EXCEL_SHEET:
        print(sheetname, "generating...", end="", flush=True)
        ws = wb[sheetname]
        mainboxlist.extend(parse(ws))
        print("Success")
    
    
    with open(args.jsonoutput, 'w') as file:
        json.dump(mainboxlist, file)

if __name__ == '__main__':
    main()
